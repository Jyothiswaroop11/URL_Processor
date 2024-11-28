import os
import shutil
import sys
import json
import time
import logging
from logging.handlers import RotatingFileHandler
from tkinter.constants import CURRENT
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse
import traceback
from tqdm import tqdm
import colorama
from colorama import Fore, Style
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, WebDriverException
from openpyxl.styles import PatternFill
import base64
from report_handler import ReportHandler

# Global status variables - STATUS
STATUS_SUCCESS = "Success"
STATUS_FAILED = "Failed"
STATUS_WARNING = "Warning"
STATUS_SKIP = "Skip"

# Level Variables - DISPLAY in LOGS
LEVEL_PASS = "PASS"
LEVEL_FAIL = "FAIL"
LEVEL_WARNING = "WARNING"
LEVEL_SKIP = "SKIP"
LEVEL_ERROR = "ERROR"
LEVEL_INFO = "INFO"

class URLValidator:
    # Define blocked URLs as class variables
    BLOCKED_URLS = [
        "r1az1.ztg.gso.adp.com",
        "r1az2.ztg.gso.adp.com"
    ]

    @staticmethod
    def load_config():
        """Load configuration from config.json"""
        try:
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config_path = os.path.join(project_root, 'config.json')

            if not os.path.exists(config_path):
                raise FileNotFoundError(f"Configuration file not found: {config_path}")

            with open(config_path, 'r') as f:
                config = json.load(f)

            logging.info("Configuration loaded successfully")
            return config
        except Exception as e:
            logging.error(f"Error loading configuration: {str(e)}")
            raise

    @staticmethod
    def get_project_paths():
        """Get project paths"""
        try:
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config = URLValidator.load_config()

            paths = {
                "project_root": project_root,
                "excel_path": os.path.join(project_root, config.get("excel_path", "resources/links.xlsx")),
                "chrome_driver": os.path.join(project_root,
                                              config.get("chrome_driver_path", "resources/drivers/chromedriver.exe")),
                "reports": os.path.join(project_root, "reports", "validation-reports"),
                "logs": os.path.join(project_root, "logs"),
                "logs_current": os.path.join(project_root, "logs", "Current Logs"),
                "logs_backup": os.path.join(project_root, "logs", "Backup Logs"),
                "output_excel": os.path.join(project_root, "reports", "excel-reports"),
                "backup_excel": os.path.join(project_root, "reports", "backup-excel")
            }

            logging.info("Project paths initialized")
            return paths
        except Exception as e:
            logging.error(f"Error getting project paths: {str(e)}")
            raise

    @staticmethod
    def setup_logging():
        """Initialize logging configuration"""
        try:
            paths = URLValidator.get_project_paths()
            os.makedirs(paths["logs_current"], exist_ok=True)
            os.makedirs(paths["logs_backup"], exist_ok=True)
            current_log_file = os.path.join(paths["logs_current"], "Current-Validation-Logs.log")

            if os.path.exists(current_log_file):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_filename = f"Previous-Validation-Logs-{timestamp}.log"
                backup_path = os.path.join(paths["logs_backup"], backup_filename)

                try:
                    shutil.copy2(current_log_file, backup_path)
                    open(current_log_file, 'w').close()
                except Exception as e:
                    print(f"Error backing up log file: {str(e)}")

            root_logger = logging.getLogger()
            root_logger.setLevel(logging.DEBUG)
            root_logger.handlers.clear()

            file_formatter = logging.Formatter(
                '%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            console_formatter = logging.Formatter(
                '%(asctime)s - %(levelname)s - %(message)s',
                datefmt='%H:%M:%S'
            )

            file_handler = RotatingFileHandler(
                current_log_file,
                maxBytes=10 * 1024 * 1024,
                backupCount=5,
                encoding='utf-8'
            )
            file_handler.setLevel(logging.DEBUG)
            file_handler.setFormatter(file_formatter)
            root_logger.addHandler(file_handler)

            console_handler = logging.StreamHandler(sys.stdout)
            console_handler.setLevel(logging.INFO)
            console_handler.setFormatter(console_formatter)
            root_logger.addHandler(console_handler)

            logging.info("Logging setup completed")
            return current_log_file

        except Exception as e:
            print(f"Error setting up logging: {str(e)}")
            raise

    @staticmethod
    def ensure_directories():
        """Create necessary directories"""
        try:
            paths = URLValidator.get_project_paths()
            for path in paths.values():
                if path and not os.path.exists(path):
                    os.makedirs(path)
                    logging.info(f"Created directory: {path}")

            os.makedirs(paths["logs_current"], exist_ok=True)
            os.makedirs(paths["logs_backup"], exist_ok=True)
            logging.info("Directory structure verified")
        except Exception as e:
            logging.error(f"Error creating directories: {str(e)}")
            raise

    @staticmethod
    def create_web_driver():
        """Create and configure Chrome WebDriver"""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--ignore-certificate-errors")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--remote-allow-origins=*")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-popup-blocking")
            chrome_options.set_capability("acceptInsecureCerts", True)

            os.environ["no_proxy"] = "*"
            os.environ["HTTP_PROXY"] = ""
            os.environ["HTTPS_PROXY"] = ""

            paths = URLValidator.get_project_paths()
            config = URLValidator.load_config()

            if not os.path.exists(paths["chrome_driver"]):
                raise FileNotFoundError(f"ChromeDriver not found at: {paths['chrome_driver']}")

            service = Service(executable_path=paths["chrome_driver"])
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(config.get("page_load_timeout", 60))
            driver.set_window_size(1920, 1080)

            logging.info("WebDriver created successfully")
            return driver

        except Exception as e:
            logging.error(f"Error creating WebDriver: {str(e)}")
            raise

    # @staticmethod
    # def process_url(url, row_number, config):
    #     """Process a single URL and return results"""
    #     start_time = time.time()
    #     driver = None
    #     test_logs = []
    #
    #     def log_step(level, message, symbol="●"):
    #         timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    #
    #         if "Launching URL Is ==>" in message:
    #             if result.get('status') == STATUS_SUCCESS:
    #                 level = LEVEL_PASS
    #             elif result.get('status') == STATUS_FAILED:
    #                 level = LEVEL_FAIL
    #             elif result.get('status') == STATUS_WARNING:
    #                 level = LEVEL_WARNING
    #             elif result.get('status') == STATUS_SKIP:
    #                 level = LEVEL_SKIP
    #
    #         log_entry = {
    #             'timestamp': timestamp,
    #             'level': level,
    #             'message': message
    #         }
    #         test_logs.append(log_entry)
    #
    #         if level in [LEVEL_ERROR, LEVEL_FAIL, STATUS_FAILED]:
    #             print(f"{Fore.RED}✗ [{timestamp}] {message}{Style.RESET_ALL}")
    #         elif level in [LEVEL_WARNING, STATUS_WARNING]:
    #             print(f"{Fore.YELLOW}⚠ [{timestamp}] {message}{Style.RESET_ALL}")
    #         elif level in [STATUS_SUCCESS, LEVEL_PASS]:
    #             print(f"{Fore.GREEN}✓ [{timestamp}] {message}{Style.RESET_ALL}")
    #         elif level in [LEVEL_SKIP, STATUS_SKIP]:
    #             print(f"{Fore.BLUE}○ [{timestamp}] {message}{Style.RESET_ALL}")
    #         else:
    #             print(f"{Fore.CYAN}{symbol} [{timestamp}] {message}{Style.RESET_ALL}")
    #
    #     def capture_screenshot(driver, attempt_number=""):
    #         """Helper function to capture screenshot with logging"""
    #         try:
    #             time.sleep(2)
    #             driver.execute_script("document.body.style.overflow = 'hidden';")
    #             screenshot = driver.get_screenshot_as_png()
    #             driver.execute_script("document.body.style.overflow = '';")
    #             if screenshot:
    #                 return base64.b64encode(screenshot).decode('utf-8')
    #             log_step('WARNING', f"Screenshot capture returned no data{attempt_number}")
    #             return None
    #         except Exception as e:
    #             log_step('WARNING', f"Screenshot capture failed{attempt_number}: {str(e)}")
    #             return None
    #
    #     result = {
    #         'url': url,
    #         'formatted_url': URLValidator.format_url(url),
    #         'status': None,
    #         'category': 'ERROR',
    #         'load_time': 0,
    #         'screenshot_base64': None,
    #         'error': None,
    #         'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    #         'steps': [],
    #         'test_logs': []
    #     }
    #
    #     try:
    #         formatted_url = URLValidator.format_url(url)
    #         driver = URLValidator.create_web_driver()
    #
    #         try:
    #             if any(blocked_url in url.lower() for blocked_url in URLValidator.BLOCKED_URLS):
    #                 try:
    #                     driver.get(formatted_url)
    #                 except:
    #                     pass
    #
    #                 screenshot_base64 = capture_screenshot(driver)
    #                 if screenshot_base64:
    #                     result['screenshot_base64'] = screenshot_base64
    #
    #                 end_time = time.time()
    #                 load_time = (end_time - start_time) * 1000
    #
    #                 log_step(STATUS_WARNING, f"Launching URL Is ==> {url}")
    #                 log_step(STATUS_WARNING, "Page Is Blocked By PaloAlto")
    #                 log_step('INFO', f"Time Taken ==> {load_time:.2f} Milliseconds")
    #
    #                 result.update({
    #                     'status': STATUS_WARNING,
    #                     'error': "Page Is Blocked By PaloAlto",
    #                     'category': 'WARNING',
    #                     'load_time': round(load_time, 2),
    #                     'test_logs': test_logs
    #                 })
    #                 return result
    #
    #             log_step('INFO', f"Launching URL Is ==> {formatted_url}")
    #             driver.get(formatted_url)
    #             driver_end_time = time.time()
    #             load_time = (driver_end_time - start_time) * 1000
    #
    #             screenshot_base64 = capture_screenshot(driver)
    #             if screenshot_base64:
    #                 result['screenshot_base64'] = screenshot_base64
    #
    #             if not URLValidator.is_valid_url(url):
    #                 end_time = time.time()
    #                 load_time = (end_time - start_time) * 1000
    #                 result.update({
    #                     'status': STATUS_FAILED,
    #                     'error': "Invalid URL format",
    #                     'category': 'ERROR',
    #                     'test_logs': test_logs,
    #                     'load_time': round(load_time, 2)
    #                 })
    #                 return result
    #
    #             page_title = driver.title
    #             redirected_url = driver.current_url
    #             load_time = (time.time() - start_time) * 1000
    #
    #             log_step('INFO', f"Title Of The Page Is ==> {page_title}")
    #             log_step('INFO', f"Redirected URL Is ==> {redirected_url}")
    #             log_step('INFO', f"Time Taken to Launch Application - {formatted_url} ==> {load_time:.2f} Milliseconds")
    #
    #             status_result = URLValidator.check_url_status(driver)
    #             status_result['load_time'] = round(load_time, 2)
    #             status_result['screenshot_base64'] = result.get('screenshot_base64')
    #             result.update(status_result)
    #
    #         except Exception as e:
    #             screenshot_base64 = capture_screenshot(driver)
    #             if screenshot_base64:
    #                 result['screenshot_base64'] = screenshot_base64
    #
    #             end_time = time.time()
    #             load_time = (end_time - start_time) * 1000
    #             result.update({
    #                 'status': STATUS_FAILED,
    #                 'error': str(e),
    #                 'category': 'ERROR',
    #                 'load_time': round(load_time, 2)
    #             })
    #
    #     except Exception as e:
    #         end_time = time.time()
    #         load_time = (end_time - start_time) * 1000
    #         result.update({
    #             'status': STATUS_FAILED,
    #             'error': str(e),
    #             'category': 'ERROR',
    #             'load_time': round(load_time, 2)
    #         })
    #
    #     finally:
    #         if result['load_time'] == 0:
    #             end_time = time.time()
    #             result['load_time'] = round((end_time - start_time) * 1000, 2)
    #
    #         result['test_logs'] = test_logs
    #         if driver:
    #             try:
    #                 driver.quit()
    #             except Exception:
    #                 pass
    #
    #         return result

    @staticmethod
    def is_valid_url(url):
        """Validate URL format"""
        try:
            result = urlparse(URLValidator.format_url(url))
            return all([result.scheme, result.netloc])
        except:
            return False

    @staticmethod
    def format_url(url):
        """Format URL with proper protocol"""
        if not url:
            return ""
        url = url.strip()
        if not url.startswith(("https://", "http://")):
            return f"https://{url}"
        return url

    # @staticmethod
    # def check_url_status(driver):
    #     """Check URL status"""
    #     try:
    #         current_url = driver.current_url.lower()
    #         page_title = driver.title if driver.title else ""
    #
    #         navigation_start = driver.execute_script("return window.performance.timing.navigationStart")
    #         response_end = driver.execute_script("return window.performance.timing.responseEnd")
    #         duration = response_end - navigation_start
    #
    #         # Check for blocked URLs
    #         if any(blocked_url in current_url for blocked_url in URLValidator.BLOCKED_URLS):
    #             return ReportHandler.format_validation_result(
    #                 url=current_url,
    #                 status='Warning',
    #                 title=page_title,
    #                 redirected_url=current_url,
    #                 duration=duration,
    #                 error='Page Is Blocked By PaloAlto',
    #                 category='WARNING'
    #             )
    #         else:
    #             return ReportHandler.format_validation_result(
    #                 url=current_url,
    #                 status='Success',
    #                 title=page_title,
    #                 redirected_url=current_url,
    #                 duration=duration,
    #                 error=None,
    #                 category='SUCCESS'
    #             )
    #
    #     except Exception as e:
    #         duration = time.time() * 1000
    #         logging.error(f"Error in check_url_status: {str(e)}")
    #         return ReportHandler.format_validation_result(
    #             url=current_url if 'current_url' in locals() else "Unknown URL",
    #             status='Failed',
    #             title="Error",
    #             redirected_url="",
    #             #duration=0,
    #             duration=duration,
    #             error=str(e),
    #             category='SYSTEM_ERROR'
    #         )

    @staticmethod
    def process_url(url, row_number, config):
        """Process a single URL and return results"""
        start_time = time.time()
        driver = None
        test_logs = []

        def log_step(level, message, symbol="●"):
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            if "Launching URL Is ==>" in message:
                if result.get('status') == STATUS_SUCCESS:
                    level = LEVEL_PASS
                elif result.get('status') == STATUS_FAILED:
                    level = LEVEL_FAIL
                elif result.get('status') == STATUS_WARNING:
                    level = LEVEL_WARNING
                elif result.get('status') == STATUS_SKIP:
                    level = LEVEL_SKIP

            log_entry = {
                'timestamp': timestamp,
                'level': level,
                'message': message
            }
            test_logs.append(log_entry)

            if level in [LEVEL_ERROR, LEVEL_FAIL, STATUS_FAILED]:
                print(f"{Fore.RED}✗ [{timestamp}] {message}{Style.RESET_ALL}")
            elif level in [LEVEL_WARNING, STATUS_WARNING]:
                print(f"{Fore.YELLOW}⚠ [{timestamp}] {message}{Style.RESET_ALL}")
            elif level in [STATUS_SUCCESS, LEVEL_PASS]:
                print(f"{Fore.GREEN}✓ [{timestamp}] {message}{Style.RESET_ALL}")
            elif level in [LEVEL_SKIP, STATUS_SKIP]:
                print(f"{Fore.BLUE}○ [{timestamp}] {message}{Style.RESET_ALL}")
            else:
                print(f"{Fore.CYAN}{symbol} [{timestamp}] {message}{Style.RESET_ALL}")

        def capture_screenshot(driver, attempt_number=""):
            """Helper function to capture screenshot with logging"""
            try:
                time.sleep(2)
                driver.execute_script("document.body.style.overflow = 'hidden';")
                screenshot = driver.get_screenshot_as_png()
                driver.execute_script("document.body.style.overflow = '';")
                if screenshot:
                    return base64.b64encode(screenshot).decode('utf-8')
                log_step('WARNING', f"Screenshot capture returned no data{attempt_number}")
                return None
            except Exception as e:
                log_step('WARNING', f"Screenshot capture failed{attempt_number}: {str(e)}")
                return None

        result = {
            'url': url,
            'formatted_url': URLValidator.format_url(url),
            'status': None,
            'category': 'ERROR',
            'load_time': 0,
            'screenshot_base64': None,
            'error': None,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'steps': [],
            'test_logs': []
        }

        try:
            formatted_url = URLValidator.format_url(url)
            driver = URLValidator.create_web_driver()

            try:
                if any(blocked_url in url.lower() for blocked_url in URLValidator.BLOCKED_URLS):
                    try:
                        driver.get(formatted_url)
                    except:
                        pass

                    screenshot_base64 = capture_screenshot(driver)
                    if screenshot_base64:
                        result['screenshot_base64'] = screenshot_base64

                    end_time = time.time()
                    load_time = (end_time - start_time) * 1000

                    log_step(STATUS_WARNING, f"Launching URL Is ==> {url}")
                    log_step(STATUS_WARNING, "Page Is Blocked By PaloAlto")
                    log_step('INFO', f"Time Taken ==> {load_time:.2f} Milliseconds")

                    result.update({
                        'status': STATUS_WARNING,
                        'error': "Page Is Blocked By PaloAlto",
                        'category': 'WARNING',
                        'load_time': round(load_time, 2),
                        'test_logs': test_logs
                    })
                    return result

                log_step('INFO', f"Launching URL Is ==> {formatted_url}")
                driver.get(formatted_url)
                driver_end_time = time.time()
                load_time = (driver_end_time - start_time) * 1000

                screenshot_base64 = capture_screenshot(driver)
                if screenshot_base64:
                    result['screenshot_base64'] = screenshot_base64

                # For invalid URL case
                if not URLValidator.is_valid_url(url):
                    end_time = time.time()
                    load_time = (end_time - start_time) * 1000
                    result.update({
                        'status': STATUS_FAILED,
                        'error': "Invalid URL format",
                        'category': 'ERROR',
                        'test_logs': test_logs,
                        'load_time': round(load_time, 2)
                    })
                    log_step('INFO', f"Time Taken ==> {load_time:.2f} Milliseconds")
                    return result

                page_title = driver.title
                redirected_url = driver.current_url
                load_time = (time.time() - start_time) * 1000

                log_step('INFO', f"Title Of The Page Is ==> {page_title}")
                log_step('INFO', f"Redirected URL Is ==> {redirected_url}")
                log_step('INFO', f"Time Taken to Launch Application - {formatted_url} ==> {load_time:.2f} Milliseconds")

                status_result = URLValidator.check_url_status(driver)
                status_result['load_time'] = round(load_time, 2)
                status_result['screenshot_base64'] = result.get('screenshot_base64')
                result.update(status_result)

            # For exception case inside try block
            except Exception as e:
                screenshot_base64 = capture_screenshot(driver)
                if screenshot_base64:
                    result['screenshot_base64'] = screenshot_base64

                end_time = time.time()
                load_time = (end_time - start_time) * 1000
                result.update({
                    'status': STATUS_FAILED,
                    'error': str(e),
                    'category': 'ERROR',
                    'load_time': round(load_time, 2)
                })
                log_step('INFO', f"Time Taken ==> {load_time:.2f} Milliseconds")
                return result

        # For outer exception case
        except Exception as e:
            end_time = time.time()
            load_time = (end_time - start_time) * 1000
            result.update({
                'status': STATUS_FAILED,
                'error': str(e),
                'category': 'ERROR',
                'load_time': round(load_time, 2)
            })
            log_step('INFO', f"Time Taken ==> {load_time:.2f} Milliseconds")  # Add this line
            return result

        finally:
            if result['load_time'] == 0:
                end_time = time.time()
                result['load_time'] = round((end_time - start_time) * 1000, 2)

            result['test_logs'] = test_logs
            if driver:
                try:
                    driver.quit()
                except Exception:
                    pass

            return result

    @staticmethod
    def check_url_status(driver):
        """Check URL status"""
        try:
            current_url = driver.current_url.lower()
            page_title = driver.title if driver.title else ""
            page_source = driver.page_source.lower()

            navigation_start = driver.execute_script("return window.performance.timing.navigationStart")
            response_end = driver.execute_script("return window.performance.timing.responseEnd")
            duration = response_end - navigation_start

            if any(blocked_url in current_url for blocked_url in URLValidator.BLOCKED_URLS):
                return ReportHandler.format_validation_result(
                    url=current_url,
                    status=STATUS_WARNING,
                    title=page_title,
                    redirected_url=current_url,
                    duration=duration,
                    error='Page Is Blocked By PaloAlto',
                    category='BLOCKED-PAGE'
                )
            elif "this page isn't working" in page_source:
                return ReportHandler.format_validation_result(
                    url=current_url,
                    status=STATUS_FAILED,
                    title=page_title,
                    redirected_url=current_url,
                    duration=duration,
                    error='Page is failing',
                    category='ERROR'
                )
            elif "web - access blocked" in page_source:
                return ReportHandler.format_validation_result(
                    url=current_url,
                    status=STATUS_SKIP,
                    title=page_title,
                    redirected_url=current_url,
                    duration=duration,
                    error='Web Access Blocked',
                    category='SKIP'
                )
            else:
                return ReportHandler.format_validation_result(
                    url=current_url,
                    status=STATUS_SUCCESS,
                    title=page_title,
                    redirected_url=current_url,
                    duration=duration,
                    error=None,
                    category='PASSED'
                )

        except Exception as e:
            logging.error(f"Error in check_url_status: {str(e)}")
            return ReportHandler.format_validation_result(
                url=current_url if 'current_url' in locals() else "Unknown URL",
                status=STATUS_FAILED,
                title="Error",
                redirected_url="",
                duration=0,
                error=str(e),
                category='SYSTEM_ERROR'
            )

    @staticmethod
    def read_excel_urls(excel_path, sheet_name):
        """Read URLs from Excel file"""
        try:
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel file not found: {excel_path}")

            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            if 'URL' not in df.columns:
                raise ValueError("Excel file must contain a 'URL' column")

            urls = df['URL'].dropna().str.strip().tolist()
            logging.info(f"Read {len(urls)} URLs from Excel")
            return [(url, idx + 2) for idx, url in enumerate(urls)]

        except Exception as e:
            logging.error(f"Error reading Excel file: {str(e)}")
            raise

    @staticmethod
    def backup_previous_excel():
        """Backup previous Excel report"""
        try:
            paths = URLValidator.get_project_paths()
            excel_dir = paths["output_excel"]
            backup_dir = paths["backup_excel"]

            os.makedirs(backup_dir, exist_ok=True)

            if not os.path.exists(excel_dir) or not os.listdir(excel_dir):
                return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            for filename in os.listdir(excel_dir):
                if filename.endswith('.xlsx'):
                    src_path = os.path.join(excel_dir, filename)
                    backup_filename = f"Previous-{filename.replace('.xlsx', '')}_{timestamp}.xlsx"
                    dst_path = os.path.join(backup_dir, backup_filename)

                    try:
                        if os.path.exists(src_path) and os.access(src_path, os.R_OK):
                            if os.path.exists(dst_path):
                                os.remove(dst_path)
                            shutil.copy2(src_path, dst_path)
                            os.remove(src_path)
                            logging.info(f"Backed up Excel report: {dst_path}")
                        else:
                            logging.warning(f"Source file not accessible or missing: {src_path}")

                    except Exception as e:
                        logging.error(f"Error backing up {src_path}: {str(e)}")
                        continue

        except Exception as e:
            logging.error(f"Error during Excel backup: {str(e)}")

    @staticmethod
    def generate_excel_report(results):
        """Generate Excel report from results"""
        try:
            paths = URLValidator.get_project_paths()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            URLValidator.backup_previous_excel()

            data = []
            for result in results:
                load_time = result.get('load_time', 0)
                # Convert to float and ensure it's not None
                if load_time is None:
                    load_time = 0.0
                elif isinstance(load_time, str):
                    try:
                        load_time = float(load_time)
                    except (ValueError, TypeError):
                        load_time = 0.0

                row = {
                    'URL': result.get('formatted_url', result['url']),
                    'Status': result['status'],
                    'Load Time (ms)': f"{float(load_time):.2f}",  # Format as string with 2 decimal places
                    #'Error/Warning': result.get('error', '')
                }
                data.append(row)

            df = pd.DataFrame(data)

            # Convert load time to numeric and format
            df['Load Time (ms)'] = pd.to_numeric(df['Load Time (ms)'], errors='coerce').fillna(0)

            os.makedirs(paths["output_excel"], exist_ok=True)
            output_path = os.path.join(paths["output_excel"], "URL-Validation-Report.xlsx")

            try:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']

                    # Format the Load Time column
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=3)  # Load Time column
                        cell.number_format = '0.00'  # Force 2 decimal places

                    # Define styles for different statuses
                    success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    skip_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

                    # Apply conditional formatting
                    for row in range(2, len(df) + 2):  # Skip header row
                        status_cell = worksheet.cell(row=row, column=2)  # Status column
                        if status_cell.value == 'Success':
                            status_cell.fill = success_fill
                        elif status_cell.value == 'Warning':
                            status_cell.fill = warning_fill
                        elif status_cell.value == 'Failed':
                            status_cell.fill = fail_fill
                        elif status_cell.value == 'Skip':
                            status_cell.fill = skip_fill

            except PermissionError:
                alt_output_path = os.path.join(paths["output_excel"], f"URL-Validation-Report_{timestamp}.xlsx")
                df.to_excel(alt_output_path, index=False, engine='openpyxl')
                output_path = alt_output_path
                logging.warning(f"Could not save to default filename, saved as: {alt_output_path}")

            logging.info(f"Excel report generated: {output_path}")
            return output_path

        except Exception as e:
            logging.error(f"Error generating Excel report: {str(e)}")
            raise


class ValidationRunner:
    def __init__(self):
        """Initialize the validation runner"""
        colorama.init(autoreset=True)
        self.start_time = datetime.now()
        self.log_path = URLValidator.setup_logging()
        self.setup_console_output()
        print(f"{Fore.YELLOW}--------------------------")
        logging.info("ValidationRunner initialized")
        logging.info(f"Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logging.info(f"Log File: {self.log_path}")

    def setup_console_output(self):
        """Setup console output"""
        os.system('cls' if os.name == 'nt' else 'clear')
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        formatter = logging.Formatter(
            f'{Fore.YELLOW}%(asctime)s{Style.RESET_ALL} - '
            f'{Fore.CYAN}%(levelname)s{Style.RESET_ALL} - '
            f'%(message)s',
            datefmt='%H:%M:%S'
        )
        console_handler.setFormatter(formatter)

        root_logger = logging.getLogger()
        for handler in root_logger.handlers[:]:
            if isinstance(handler, logging.StreamHandler) and handler.stream == sys.stdout:
                root_logger.removeHandler(handler)
        root_logger.addHandler(console_handler)

    def print_header(self):
        """Print header information"""
        print(f"\n{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}URL Validation Process{Style.RESET_ALL}".center(80))
        print(f"{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}\n")
        print(f"{Fore.YELLOW}Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Log File: {self.log_path}{Style.RESET_ALL}\n")

    def print_summary(self, results):
        """Print validation summary"""
        end_time = datetime.now()
        duration = end_time - self.start_time

        success_count = sum(1 for r in results if r['status'] == STATUS_SUCCESS)
        warning_count = sum(1 for r in results if r['status'] == STATUS_WARNING)
        skip_count = sum(1 for r in results if r['status'] == STATUS_SKIP)
        fail_count = sum(1 for r in results if r['status'] == STATUS_FAILED)

        total_excluding_skipped = len(results) - skip_count
        success_rate = (success_count / total_excluding_skipped * 100) if total_excluding_skipped > 0 else 0

        print(f"\n{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}Validation Summary{Style.RESET_ALL}".center(80))
        print(f"{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}\n")

        print(f"{Fore.YELLOW}Duration: {duration}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Total URLs Processed: {len(results)}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}✓ Successful: {success_count}{Style.RESET_ALL}")
        print(f"{Fore.RED}✗ Failed: {fail_count}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}⚠ Warnings: {warning_count}{Style.RESET_ALL}")
        print(f"{Fore.BLUE}○ Skipped: {skip_count}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Success Rate: {success_rate:.2f}%{Style.RESET_ALL}\n")
        print(f"{Fore.CYAN}{'-' * 40}{Style.RESET_ALL}")

        print(f"{Fore.GREEN}Time Duration{Style.RESET_ALL}".center(40))
        print(f"{Fore.CYAN}{'-' * 40}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Start Time: {self.start_time.strftime('%H:%M:%S')}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}End Time: {self.end_time.strftime('%H:%M:%S')}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Duration: {duration}{Style.RESET_ALL}")
        print(f"{Fore.CYAN}{'=' * 40}{Style.RESET_ALL}")

        # if fail_count > 0:
        #     print(f"{Fore.RED}Failed URLs:{Style.RESET_ALL}")
        #     for result in results:
        #         if result['status'] == STATUS_FAILED:
        #             print(f"{Fore.RED}✗ {result['url']}: {result.get('error', 'Unknown error')}{Style.RESET_ALL}")
        #
        # if warning_count > 0:
        #     print(f"\n{Fore.YELLOW}Warning URLs:{Style.RESET_ALL}")
        #     for result in results:
        #         if result['status'] == STATUS_WARNING:
        #             print(
        #                 f"{Fore.YELLOW}⚠ {result['url']}: {result.get('error', 'Performance/Security Warning')}{Style.RESET_ALL}")
        #
        # if skip_count > 0:
        #     print(f"\n{Fore.BLUE}Skipped URLs:{Style.RESET_ALL}")
        #     for result in results:
        #         if result['status'] == STATUS_SKIP:
        #             print(f"{Fore.BLUE}○ {result['url']}: {result.get('error', 'Unknown error')}{Style.RESET_ALL}")

    def validate_urls(self):
        """Run the URL validation process"""
        try:
            self.print_header()
            config = URLValidator.load_config()
            paths = URLValidator.get_project_paths()

            print(f"\n{Fore.CYAN}Reading URLs from Excel...{Style.RESET_ALL}")
            urls = URLValidator.read_excel_urls(paths["excel_path"], config["sheet_name"])

            if not urls:
                print(f"\n{Fore.RED}✗ Error: No URLs found in Excel file{Style.RESET_ALL}")
                return

            total_urls = len(urls)
            print(f"\n{Fore.GREEN}✓ Found {total_urls} URLs to process{Style.RESET_ALL}")
            print(f"\n{Fore.GREEN}-----------------------------")

            results = []
            with tqdm(total=total_urls, desc="Processing URLs",
                      bar_format="{l_bar}%s{bar}%s{r_bar}" % (Fore.GREEN, Style.RESET_ALL)) as pbar:

                print("newline after the progress bar is initialized")

                for i, (url, row_number) in enumerate(urls, 1):
                    try:
                        result = URLValidator.process_url(url, row_number, config)
                        results.append(result)

                        status_symbol = {
                            'Success': ('✓', Fore.GREEN),
                            'Warning': ('⚠', Fore.YELLOW),
                            'Skip': ('○', Fore.BLUE),
                            'Failed': ('✗', Fore.RED)
                        }.get(result['status'], ('✗', Fore.RED))

                        pbar.set_postfix_str(
                            f"{status_symbol[1]}{status_symbol[0]} {result['status']}{Style.RESET_ALL}",
                            refresh=True
                        )
                        pbar.update(1)

                        if i < total_urls:
                            time.sleep(config.get("wait_between_urls", 2))

                    except Exception as e:
                        print(f"\n{Fore.RED}✗ Error processing URL {url}: {str(e)}{Style.RESET_ALL}")
                        results.append({
                            'url': url,
                            'status': 'Failed',
                            'error': str(e),
                            'load_time': 0,
                            'steps': []
                        })
                        pbar.update(1)

                    print(f"\n{Fore.GREEN}=========================")

            print(f"\n{Fore.CYAN}Generating reports...{Style.RESET_ALL}")

            try:
                excel_report = URLValidator.generate_excel_report(results)
                print(f"{Fore.GREEN}✓ Excel report generated: {excel_report}{Style.RESET_ALL}")

                html_report = ReportHandler.generate_detailed_html_report(results, paths["reports"])
                print(f"{Fore.GREEN}✓ HTML report generated: {html_report}{Style.RESET_ALL}")

                self.end_time = datetime.now()
                print(f"End Time: {self.end_time.strftime('%Y-%m-%d %H:%M:%S')}")
                print(f"{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}")

            except Exception as e:
                print(f"{Fore.RED}✗ Error generating reports: {str(e)}{Style.RESET_ALL}")

            self.print_summary(results)

        except Exception as e:
            self.end_time = datetime.now()
            print(f"\n{Fore.RED}✗ Validation process failed: {str(e)}{Style.RESET_ALL}")
            print(f"\n{Fore.RED}Stack trace:{Style.RESET_ALL}")
            print(traceback.format_exc())

        finally:
            colorama.deinit()


def main():
    """Main entry point"""
    try:
        print(f"\n{Fore.CYAN}Starting URL validation process...{Style.RESET_ALL}")
        runner = ValidationRunner()
        runner.validate_urls()
        print(f"\n{Fore.GREEN}✓ URL validation process completed successfully{Style.RESET_ALL}")

    except Exception as e:
        print(f"\n{Fore.RED}✗ Fatal error in main process: {str(e)}{Style.RESET_ALL}")
        print(f"\n{Fore.RED}Stack trace:{Style.RESET_ALL}")
        print(traceback.format_exc())
        sys.exit(1)

    finally:
        print(f"\n{Fore.YELLOW}Process finished. Check reports for detailed results.{Style.RESET_ALL}")


if __name__ == "__main__":
    main()
