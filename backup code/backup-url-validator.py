# import os
# import shutil
# import sys
# import json
# import time
# import logging
# from logging.handlers import RotatingFileHandler
# import pandas as pd
# from datetime import datetime
# from urllib.parse import urlparse
# import traceback
# from tqdm import tqdm
# import colorama
# from colorama import Fore, Style
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import TimeoutException, WebDriverException
# from openpyxl.styles import PatternFill
# import base64
# from report_handler import ReportHandler
#
# class URLValidator:
#     @staticmethod
#     def load_config():
#         """Load configuration from config.json"""
#         try:
#             project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#             config_path = os.path.join(project_root, 'config.json')
#
#             if not os.path.exists(config_path):
#                 raise FileNotFoundError(f"Configuration file not found: {config_path}")
#
#             with open(config_path, 'r') as f:
#                 config = json.load(f)
#
#             logging.info("Configuration loaded successfully")
#             return config
#         except Exception as e:
#             logging.error(f"Error loading configuration: {str(e)}")
#             raise
#
#     @staticmethod
#     def get_project_paths():
#         """Get project paths"""
#         try:
#             project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
#             config = URLValidator.load_config()
#
#             paths = {
#                 "project_root": project_root,
#                 "excel_path": os.path.join(project_root, config.get("excel_path", "resources/links.xlsx")),
#                 "chrome_driver": os.path.join(project_root,
#                                             config.get("chrome_driver_path", "resources/drivers/chromedriver.exe")),
#                 "reports": os.path.join(project_root, "reports", "validation-reports"),
#                 "logs": os.path.join(project_root, "logs"),
#                 "logs_current": os.path.join(project_root, "logs", "Current Logs"),
#                 "logs_backup": os.path.join(project_root, "logs", "Backup Logs"),
#                 "output_excel": os.path.join(project_root, "reports", "excel-reports"),
#                 "backup_excel": os.path.join(project_root, "reports", "backup-excel")
#             }
#
#             logging.info("Project paths initialized")
#             return paths
#         except Exception as e:
#             logging.error(f"Error getting project paths: {str(e)}")
#             raise
#
#     @staticmethod
#     def setup_logging():
#         """Initialize logging configuration"""
#         try:
#             paths = URLValidator.get_project_paths()
#
#             os.makedirs(paths["logs_current"], exist_ok=True)
#             os.makedirs(paths["logs_backup"], exist_ok=True)
#
#             current_log_file = os.path.join(paths["logs_current"], "Current-Validation-Logs.log")
#
#             if os.path.exists(current_log_file):
#                 timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#                 backup_filename = f"Previous-Validation-Logs-{timestamp}.log"
#                 backup_path = os.path.join(paths["logs_backup"], backup_filename)
#
#                 try:
#                     shutil.copy2(current_log_file, backup_path)
#                     open(current_log_file, 'w').close()
#                 except Exception as e:
#                     print(f"Error backing up log file: {str(e)}")
#
#             root_logger = logging.getLogger()
#             root_logger.setLevel(logging.DEBUG)
#             root_logger.handlers.clear()
#
#             file_formatter = logging.Formatter(
#                 '%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
#                 datefmt='%Y-%m-%d %H:%M:%S'
#             )
#             console_formatter = logging.Formatter(
#                 '%(asctime)s - %(levelname)s - %(message)s',
#                 datefmt='%H:%M:%S'
#             )
#
#             file_handler = RotatingFileHandler(
#                 current_log_file,
#                 maxBytes=10*1024*1024,
#                 backupCount=5,
#                 encoding='utf-8'
#             )
#             file_handler.setLevel(logging.DEBUG)
#             file_handler.setFormatter(file_formatter)
#             root_logger.addHandler(file_handler)
#
#             console_handler = logging.StreamHandler(sys.stdout)
#             console_handler.setLevel(logging.INFO)
#             console_handler.setFormatter(console_formatter)
#             root_logger.addHandler(console_handler)
#
#             logging.info("Logging setup completed")
#             return current_log_file
#
#         except Exception as e:
#             print(f"Error setting up logging: {str(e)}")
#             raise
#
#     @staticmethod
#     def ensure_directories():
#         """Create necessary directories"""
#         try:
#             paths = URLValidator.get_project_paths()
#             for path in paths.values():
#                 if path and not os.path.exists(path):
#                     os.makedirs(path)
#                     logging.info(f"Created directory: {path}")
#
#             os.makedirs(paths["logs_current"], exist_ok=True)
#             os.makedirs(paths["logs_backup"], exist_ok=True)
#
#             logging.info("Directory structure verified")
#         except Exception as e:
#             logging.error(f"Error creating directories: {str(e)}")
#             raise
#
#     @staticmethod
#     def create_web_driver():
#         """Create and configure Chrome WebDriver"""
#         try:
#             chrome_options = Options()
#             chrome_options.add_argument("--headless")
#             chrome_options.add_argument("--no-sandbox")
#             chrome_options.add_argument("--ignore-certificate-errors")
#             chrome_options.add_argument("--disable-dev-shm-usage")
#             chrome_options.add_argument("--remote-allow-origins=*")
#             chrome_options.add_argument("--window-size=1920,1080")
#             chrome_options.add_argument("--start-maximized")
#             chrome_options.add_argument("--disable-gpu")
#             chrome_options.add_argument("--disable-extensions")
#             chrome_options.add_argument("--disable-software-rasterizer")
#             chrome_options.add_argument("--disable-popup-blocking")
#             chrome_options.add_argument("--force-device-scale-factor=1")
#
#             chrome_options.set_capability("acceptInsecureCerts", True)
#
#             chrome_options.add_argument("--proxy-server=direct://")
#             chrome_options.add_argument("--proxy-bypass-list=*")
#             chrome_options.add_argument("--no-proxy-server")
#
#             os.environ["no_proxy"] = "*"
#             os.environ["NO_PROXY"] = "*"
#             os.environ["HTTP_PROXY"] = ""
#             os.environ["HTTPS_PROXY"] = ""
#
#             paths = URLValidator.get_project_paths()
#             config = URLValidator.load_config()
#
#             if not os.path.exists(paths["chrome_driver"]):
#                 raise FileNotFoundError(f"ChromeDriver not found at: {paths['chrome_driver']}")
#
#             service = Service(executable_path=paths["chrome_driver"])
#             driver = webdriver.Chrome(service=service, options=chrome_options)
#             driver.set_page_load_timeout(config.get("page_load_timeout", 60))
#             driver.set_window_size(1920, 1080)
#
#             logging.info("WebDriver created successfully")
#             return driver
#
#         except Exception as e:
#             logging.error(f"Error creating WebDriver: {str(e)}")
#             raise
#
#     # @staticmethod
#     # def process_url(url, row_number, config):
#     #     """Process a single URL and return results with complete functionality"""
#     #     start_time = time.time()
#     #     driver = None
#     #     test_logs = []
#     #
#     #     def log_step(level, message, symbol="●"):
#     #         """Helper function to log steps with timestamp and symbols"""
#     #         timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #         log_entry = {
#     #             'timestamp': timestamp,
#     #             'level': level,
#     #             'message': message
#     #         }
#     #         test_logs.append(log_entry)
#     #
#     #         if level == 'ERROR':
#     #             print(f"{Fore.RED}✗ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         elif level == 'WARNING':
#     #             print(f"{Fore.YELLOW}⚠ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         elif level == 'SUCCESS':
#     #             print(f"{Fore.GREEN}✓ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         else:
#     #             print(f"{Fore.CYAN}{symbol} [{timestamp}] {message}{Style.RESET_ALL}")
#     #
#     #     print(f"\n{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}")
#     #     print(f"{Fore.GREEN}Processing URL {row_number - 1}: {url}{Style.RESET_ALL}")
#     #     print(f"{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}\n")
#     #
#     #     result = {
#     #         'url': url,
#     #         'status': None,
#     #         'category': 'ERROR',
#     #         'load_time': 0,
#     #         'screenshot_base64': None,
#     #         'error': None,
#     #         'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
#     #         'steps': [],
#     #         'test_logs': []
#     #     }
#     #
#     #     try:
#     #         log_step('INFO', f"Starting validation for URL: {url}", "►")
#     #
#     #         if not URLValidator.is_valid_url(url):
#     #             log_step('ERROR', f"Invalid URL format: {url}")
#     #             result.update({
#     #                 'status': 'Failed',
#     #                 'error': "Invalid URL format",
#     #                 'category': 'ERROR',
#     #                 'test_logs': test_logs
#     #             })
#     #             return result
#     #
#     #         log_step('INFO', "Creating WebDriver instance...", "⚙")
#     #         driver = URLValidator.create_web_driver()
#     #         log_step('SUCCESS', "WebDriver created successfully")
#     #
#     #         if "ceas.sase.responses.es.oneadp.com" in url:
#     #             try:
#     #                 log_step('INFO', f"Loading blocked URL for screenshot: {url}", "►")
#     #                 driver.get(url)
#     #
#     #                 try:
#     #                     log_step('INFO', "Waiting for PaloAlto block page elements...", "⌛")
#     #                     WebDriverWait(driver, 10).until(
#     #                         lambda d: "access blocked" in d.page_source.lower() or
#     #                                   "paloalto" in d.page_source.lower() or
#     #                                   "block page" in d.page_source.lower()
#     #                     )
#     #                     log_step('SUCCESS', "PaloAlto block page detected")
#     #                 except:
#     #                     log_step('WARNING', "Timeout waiting for PaloAlto block page")
#     #
#     #                 time.sleep(5)
#     #
#     #                 log_step('INFO', "Capturing screenshot...", "📸")
#     #                 screenshot_base64 = URLValidator.capture_screenshot_base64(driver)
#     #                 if screenshot_base64:
#     #                     result['screenshot_base64'] = screenshot_base64
#     #                     log_step('SUCCESS', "Screenshot captured successfully")
#     #                 else:
#     #                     log_step('WARNING', "Screenshot capture returned no data")
#     #             except Exception as screenshot_error:
#     #                 log_step('WARNING', f"Screenshot capture failed: {str(screenshot_error)}")
#     #
#     #             log_step('INFO', "URL is blocked by PaloAlto")
#     #             result.update({
#     #                 'status': 'Skip',
#     #                 'error': 'Page Is Blocked By PaloAlto',
#     #                 'category': 'PALOALTO_BLOCK',
#     #                 'steps': [{
#     #                     'status': 'SKIP',
#     #                     'message': f'URL is blocked by PaloAlto\nURL: {url}',
#     #                     'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #                 }],
#     #                 'test_logs': test_logs
#     #             })
#     #             return result
#     #
#     #         formatted_url = URLValidator.format_url(url)
#     #         log_step('INFO', f"Formatted URL: {formatted_url}", "►")
#     #
#     #         try:
#     #             log_step('INFO', "Attempting to load URL...", "►")
#     #             driver.get(formatted_url)
#     #             log_step('SUCCESS', "Initial URL load successful")
#     #
#     #             log_step('INFO', "Checking page status...", "⌛")
#     #             status_result = URLValidator.check_url_status(driver)
#     #             result.update(status_result)
#     #             log_step('INFO', f"Page status: {status_result['status']}", "ℹ")
#     #
#     #             try:
#     #                 time.sleep(2)
#     #                 log_step('INFO', "Capturing page screenshot...", "📸")
#     #                 screenshot_base64 = URLValidator.capture_screenshot_base64(driver)
#     #                 if screenshot_base64:
#     #                     result['screenshot_base64'] = screenshot_base64
#     #                     log_step('SUCCESS', "Screenshot captured successfully")
#     #                 else:
#     #                     log_step('WARNING', "Screenshot capture returned no data")
#     #             except Exception as screenshot_error:
#     #                 log_step('WARNING', f"Screenshot capture failed: {str(screenshot_error)}")
#     #
#     #             if result['status'] != 'Skip':
#     #                 log_step('INFO', "Performing complete page load check...", "⌛")
#     #                 page_loaded = URLValidator.check_page_loaded(
#     #                     driver,
#     #                     timeout=config.get("page_load_timeout", 30),
#     #                     retries=config.get("max_retries", 2),
#     #                     retry_delay=config.get("retry_delay", 2)
#     #                 )
#     #
#     #                 load_time = (time.time() - start_time) * 1000
#     #                 result['load_time'] = round(load_time, 2)
#     #                 log_step('INFO', f"Page load time: {result['load_time']} ms", "⏱")
#     #
#     #                 if not page_loaded:
#     #                     log_step('ERROR', "Page failed to load completely")
#     #                     result['status'] = 'Failed'
#     #                     result['error'] = "Page failed to load completely"
#     #                 else:
#     #                     log_step('SUCCESS', "Page loaded successfully")
#     #
#     #         except Exception as e:
#     #             error_msg = str(e)
#     #             log_step('ERROR', f"Error during page load: {error_msg}")
#     #             if result['status'] != 'Skip':
#     #                 result.update({
#     #                     'status': 'Failed',
#     #                     'error': error_msg,
#     #                     'category': 'ERROR',
#     #                     'test_logs': test_logs
#     #                 })
#     #
#     #     except Exception as e:
#     #         error_msg = str(e)
#     #         log_step('ERROR', f"Fatal error: {error_msg}")
#     #         if result['status'] != 'Skip':
#     #             result.update({
#     #                 'status': 'Failed',
#     #                 'error': error_msg,
#     #                 'category': 'ERROR',
#     #                 'test_logs': test_logs
#     #             })
#     #
#     #     finally:
#     #         result['test_logs'] = test_logs
#     #         if driver:
#     #             try:
#     #                 if not result.get('screenshot_base64'):
#     #                     try:
#     #                         log_step('INFO', "Attempting final screenshot capture...", "📸")
#     #                         screenshot_base64 = URLValidator.capture_screenshot_base64(driver)
#     #                         if screenshot_base64:
#     #                             result['screenshot_base64'] = screenshot_base64
#     #                             log_step('SUCCESS', "Final screenshot captured successfully")
#     #                     except Exception as e:
#     #                         log_step('WARNING', f"Final screenshot capture failed: {str(e)}")
#     #
#     #                 log_step('INFO', "Closing WebDriver...", "⚙")
#     #                 driver.quit()
#     #                 log_step('SUCCESS', "WebDriver closed successfully")
#     #             except Exception as e:
#     #                 log_step('WARNING', f"Error closing WebDriver: {str(e)}")
#     #
#     #         if result['status'] is None:
#     #             result['status'] = 'Failed'
#     #             log_step('WARNING', "Status was not set, defaulting to Failed")
#     #
#     #         log_step('INFO', f"URL validation complete. Final status: {result['status']}", "✓")
#     #
#     #     return result
#
#     # @staticmethod
#     # def process_url(url, row_number, config):
#     #     """Process a single URL and return results with complete functionality"""
#     #     start_time = time.time()
#     #     driver = None
#     #     test_logs = []
#     #
#     #     def log_step(level, message, symbol="●"):
#     #         timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #         log_entry = {
#     #             'timestamp': timestamp,
#     #             'level': level,
#     #             'message': message
#     #         }
#     #         test_logs.append(log_entry)
#     #
#     #         if level == 'ERROR':
#     #             print(f"{Fore.RED}✗ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         elif level == 'WARNING':
#     #             print(f"{Fore.YELLOW}⚠ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         elif level == 'SUCCESS':
#     #             print(f"{Fore.GREEN}✓ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         else:
#     #             print(f"{Fore.CYAN}{symbol} [{timestamp}] {message}{Style.RESET_ALL}")
#     #
#     #     print(f"\n{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}")
#     #     print(f"{Fore.GREEN}Processing URL {row_number - 1}: {url}{Style.RESET_ALL}")
#     #     print(f"{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}\n")
#     #
#     #     result = {
#     #         'url': url,
#     #         'status': None,
#     #         'category': 'ERROR',
#     #         'load_time': 0,
#     #         'screenshot_base64': None,
#     #         'error': None,
#     #         'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
#     #         'steps': [],
#     #         'test_logs': []
#     #     }
#     #
#     #     try:
#     #         log_step('INFO', f"Starting validation for URL: {url}", "►")
#     #
#     #         if not URLValidator.is_valid_url(url):
#     #             log_step('ERROR', f"Invalid URL format: {url}")
#     #             result.update({
#     #                 'status': 'Failed',
#     #                 'error': "Invalid URL format",
#     #                 'category': 'ERROR',
#     #                 'test_logs': test_logs
#     #             })
#     #             return result
#     #
#     #         log_step('INFO', "Creating WebDriver instance...", "⚙")
#     #         driver = URLValidator.create_web_driver()
#     #         log_step('SUCCESS', "WebDriver created successfully")
#     #
#     #         formatted_url = URLValidator.format_url(url)
#     #         log_step('INFO', f"Launching URL Is ==> {formatted_url}", "►")
#     #
#     #         try:
#     #             log_step('INFO', "Attempting to load URL...", "►")
#     #             driver.get(formatted_url)
#     #             log_step('SUCCESS', "Initial URL load successful")
#     #
#     #             # Log page title and redirected URL
#     #             log_step('INFO', f"Title Of The Page Is ==> {driver.title}")
#     #             log_step('INFO', f"Redirected URL Is ==> {driver.current_url}")
#     #             load_time = (time.time() - start_time) * 1000
#     #             log_step('INFO', f"Time Taken to Launch Application - {formatted_url} ==> {load_time:.2f} Milliseconds")
#     #
#     #             log_step('INFO', "Checking page status...", "⌛")
#     #             status_result = URLValidator.check_url_status(driver)
#     #             result.update(status_result)
#     #             log_step('INFO', f"Page status: {status_result['status']}", "ℹ")
#     #
#     #             try:
#     #                 time.sleep(2)
#     #                 log_step('INFO', "Capturing page screenshot...", "📸")
#     #                 screenshot_base64 = URLValidator.capture_screenshot_base64(driver)
#     #                 if screenshot_base64:
#     #                     result['screenshot_base64'] = screenshot_base64
#     #                     log_step('SUCCESS', "Screenshot captured successfully")
#     #                 else:
#     #                     log_step('WARNING', "Screenshot capture returned no data")
#     #             except Exception as screenshot_error:
#     #                 log_step('WARNING', f"Screenshot capture failed: {str(screenshot_error)}")
#     #
#     #             if result['status'] != 'Skip':
#     #                 log_step('INFO', "Performing complete page load check...", "⌛")
#     #                 page_loaded = URLValidator.check_page_loaded(
#     #                     driver,
#     #                     timeout=config.get("page_load_timeout", 30),
#     #                     retries=config.get("max_retries", 2),
#     #                     retry_delay=config.get("retry_delay", 2)
#     #                 )
#     #
#     #                 result['load_time'] = round(load_time, 2)
#     #                 log_step('INFO', f"Page load time: {result['load_time']} ms", "⏱")
#     #
#     #                 if not page_loaded:
#     #                     log_step('ERROR', "Page failed to load completely")
#     #                     result['status'] = 'Failed'
#     #                     result['error'] = "Page failed to load completely"
#     #                 else:
#     #                     log_step('SUCCESS', "Page loaded successfully")
#     #
#     #         except Exception as e:
#     #             error_msg = str(e)
#     #             log_step('ERROR', f"Error during page load: {error_msg}")
#     #             if result['status'] != 'Skip':
#     #                 result.update({
#     #                     'status': 'Failed',
#     #                     'error': error_msg,
#     #                     'category': 'ERROR',
#     #                     'test_logs': test_logs
#     #                 })
#     #
#     #     except Exception as e:
#     #         error_msg = str(e)
#     #         log_step('ERROR', f"Fatal error: {error_msg}")
#     #         if result['status'] != 'Skip':
#     #             result.update({
#     #                 'status': 'Failed',
#     #                 'error': error_msg,
#     #                 'category': 'ERROR',
#     #                 'test_logs': test_logs
#     #             })
#     #
#     #     finally:
#     #         result['test_logs'] = test_logs
#     #         if driver:
#     #             try:
#     #                 if not result.get('screenshot_base64'):
#     #                     try:
#     #                         log_step('INFO', "Attempting final screenshot capture...", "📸")
#     #                         screenshot_base64 = URLValidator.capture_screenshot_base64(driver)
#     #                         if screenshot_base64:
#     #                             result['screenshot_base64'] = screenshot_base64
#     #                             log_step('SUCCESS', "Final screenshot captured successfully")
#     #                     except Exception as e:
#     #                         log_step('WARNING', f"Final screenshot capture failed: {str(e)}")
#     #
#     #                 log_step('INFO', "Closing WebDriver...", "⚙")
#     #                 driver.quit()
#     #                 log_step('SUCCESS', "WebDriver closed successfully")
#     #             except Exception as e:
#     #                 log_step('WARNING', f"Error closing WebDriver: {str(e)}")
#     #
#     #         if result['status'] is None:
#     #             result['status'] = 'Failed'
#     #             log_step('WARNING', "Status was not set, defaulting to Failed")
#     #
#     #         log_step('INFO', f"URL validation complete. Final status: {result['status']}", "✓")
#     #
#     #     return result
#
#     # @staticmethod
#     # def process_url(url, row_number, config):
#     #     """Process a single URL and return results with complete functionality"""
#     #     start_time = time.time()
#     #     driver = None
#     #     test_logs = []
#     #
#     #     def log_step(level, message, symbol="●"):
#     #         timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #
#     #         # Special handling for the launch URL message
#     #         if "Launching URL Is ==>" in message:
#     #             if result.get('status') == 'Success':
#     #                 level = 'PASS'
#     #             elif result.get('status') == 'Failed':
#     #                 level = 'FAIL'
#     #             elif result.get('status') == 'Warning':
#     #                 level = 'WARNING'
#     #             elif result.get('status') == 'Skip':
#     #                 level = 'SKIP'
#     #
#     #         log_entry = {
#     #             'timestamp': timestamp,
#     #             'level': level,
#     #             'message': message
#     #         }
#     #         test_logs.append(log_entry)
#     #
#     #         if level == 'ERROR' or level == 'FAIL' or level == 'Failed':
#     #             print(f"{Fore.RED}✗ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         elif level == 'WARNING':
#     #             print(f"{Fore.YELLOW}⚠ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         elif level == 'SUCCESS' or level == 'PASS' or level == 'Success':
#     #             print(f"{Fore.GREEN}✓ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         elif level == 'SKIP' or level == 'Skip':
#     #             print(f"{Fore.BLUE}○ [{timestamp}] {message}{Style.RESET_ALL}")
#     #         else:
#     #             print(f"{Fore.CYAN}{symbol} [{timestamp}] {message}{Style.RESET_ALL}")
#     #
#     #     result = {
#     #         'url': url,
#     #         'status': None,
#     #         'category': 'ERROR',
#     #         'load_time': 0,
#     #         'screenshot_base64': None,
#     #         'error': None,
#     #         'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
#     #         'steps': [],
#     #         'test_logs': []
#     #     }
#     #
#     #     try:
#     #         log_step('INFO', f"Starting validation for URL: {url}", "►")
#     #
#     #         # Check for PaloAlto blocked URL first
#     #         if "ceas.sase.responses.es.oneadp.com" in url:
#     #             result.update({
#     #                 'status': 'Skip',
#     #                 'error': "Access blocked by security policy",
#     #                 'category': 'PALOALTO_BLOCK',
#     #                 'test_logs': test_logs
#     #             })
#     #             formatted_url = URLValidator.format_url(url)
#     #             log_step('SKIP', f"Launching URL Is ==> {formatted_url}", "►")
#     #             return result
#     #
#     #         if not URLValidator.is_valid_url(url):
#     #             log_step('Failed', f"Invalid URL format: {url}")
#     #             result.update({
#     #                 'status': 'Failed',
#     #                 'error': "Invalid URL format",
#     #                 'category': 'ERROR',
#     #                 'test_logs': test_logs
#     #             })
#     #             return result
#     #
#     #         log_step('INFO', "Creating WebDriver instance...", "⚙")
#     #         driver = URLValidator.create_web_driver()
#     #         log_step('SUCCESS', "WebDriver created successfully")
#     #
#     #         formatted_url = URLValidator.format_url(url)
#     #         log_step('INFO', f"Launching URL Is ==> {formatted_url}", "►")
#     #
#     #         try:
#     #             log_step('INFO', "Attempting to load URL...", "►")
#     #             driver.get(formatted_url)
#     #             log_step('SUCCESS', "Initial URL load successful")
#     #
#     #             log_step('INFO', f"Title Of The Page Is ==> {driver.title}")
#     #             log_step('INFO', f"Redirected URL Is ==> {driver.current_url}")
#     #             load_time = (time.time() - start_time) * 1000
#     #             log_step('INFO', f"Time Taken to Launch Application - {formatted_url} ==> {load_time:.2f} Milliseconds")
#     #
#     #             log_step('INFO', "Checking page status...", "⌛")
#     #             status_result = URLValidator.check_url_status(driver)
#     #             result.update(status_result)
#     #             log_step('INFO', f"Page status: {status_result['status']}", "ℹ")
#     #
#     #             try:
#     #                 time.sleep(2)
#     #                 screenshot_base64 = URLValidator.capture_screenshot_base64(driver)
#     #                 if screenshot_base64:
#     #                     result['screenshot_base64'] = screenshot_base64
#     #                     log_step('SUCCESS', "Screenshot captured successfully")
#     #                 else:
#     #                     log_step('WARNING', "Screenshot capture returned no data")
#     #             except Exception as screenshot_error:
#     #                 log_step('WARNING', f"Screenshot capture failed: {str(screenshot_error)}")
#     #
#     #             if result['status'] != 'Skip':
#     #                 log_step('INFO', "Performing complete page load check...", "⌛")
#     #                 page_loaded = URLValidator.check_page_loaded(
#     #                     driver,
#     #                     timeout=config.get("page_load_timeout", 30),
#     #                     retries=config.get("max_retries", 2),
#     #                     retry_delay=config.get("retry_delay", 2)
#     #                 )
#     #
#     #                 result['load_time'] = round(load_time, 2)
#     #
#     #                 if not page_loaded:
#     #                     log_step('Failed', "Page failed to load completely")
#     #                     result['status'] = 'Failed'
#     #                     result['error'] = "Page failed to load completely"
#     #                 else:
#     #                     log_step('Success', "Page loaded successfully")
#     #                     if not result.get('status'):
#     #                         result['status'] = 'Success'
#     #
#     #         except Exception as e:
#     #             error_msg = str(e)
#     #             log_step('Failed', f"Error during page load: {error_msg}")
#     #             if result['status'] != 'Skip':
#     #                 result.update({
#     #                     'status': 'Failed',
#     #                     'error': error_msg,
#     #                     'category': 'ERROR',
#     #                     'test_logs': test_logs
#     #                 })
#     #
#     #     except Exception as e:
#     #         error_msg = str(e)
#     #         log_step('Failed', f"Fatal error: {error_msg}")
#     #         if result['status'] != 'Skip':
#     #             result.update({
#     #                 'status': 'Failed',
#     #                 'error': error_msg,
#     #                 'category': 'ERROR',
#     #                 'test_logs': test_logs
#     #             })
#     #
#     #     finally:
#     #         result['test_logs'] = test_logs
#     #         if driver:
#     #             try:
#     #                 log_step('INFO', "Closing WebDriver...", "⚙")
#     #                 driver.quit()
#     #                 log_step('SUCCESS', "WebDriver closed successfully")
#     #             except Exception as e:
#     #                 log_step('WARNING', f"Error closing WebDriver: {str(e)}")
#     #
#     #         if result['status'] is None:
#     #             result['status'] = 'Failed'
#     #             log_step('WARNING', "Status was not set, defaulting to Failed")
#     #
#     #         log_step('INFO', f"URL validation complete. Final status: {result['status']}", "✓")
#     #
#     #     return result
#
#     @staticmethod
#     def process_url(url, row_number, config):
#         """Process a single URL and return results with complete functionality"""
#         start_time = time.time()
#         driver = None
#         test_logs = []
#         screenshot_base64 = None
#
#         def log_step(level, message, symbol="●"):
#             timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#
#             # Special handling for status-specific messages
#             if "Launching URL Is ==>" in message:
#                 if result.get('status') == 'Success':
#                     level = 'PASS'
#                 elif result.get('status') == 'Failed':
#                     level = 'FAIL'
#                 elif result.get('status') == 'Warning':
#                     level = 'WARNING'
#                 elif result.get('status') == 'Skip':
#                     level = 'SKIP'
#
#             log_entry = {
#                 'timestamp': timestamp,
#                 'level': level,
#                 'message': message
#             }
#             test_logs.append(log_entry)
#
#             if level == 'ERROR' or level == 'FAIL' or level == 'Failed':
#                 print(f"{Fore.RED}✗ [{timestamp}] {message}{Style.RESET_ALL}")
#             elif level == 'WARNING' or level == 'Warning' or level == 'warning':
#                 print(f"{Fore.YELLOW}⚠ [{timestamp}] {message}{Style.RESET_ALL}")
#             elif level == 'SUCCESS' or level == 'PASS' or level == 'Success':
#                 print(f"{Fore.GREEN}✓ [{timestamp}] {message}{Style.RESET_ALL}")
#             elif level == 'SKIP' or level == 'Skip':
#                 print(f"{Fore.BLUE}○ [{timestamp}] {message}{Style.RESET_ALL}")
#             else:
#                 print(f"{Fore.CYAN}{symbol} [{timestamp}] {message}{Style.RESET_ALL}")
#
#         result = {
#             'url': url,
#             'status': None,
#             'category': 'ERROR',
#             'load_time': 0,
#             'screenshot_base64': None,
#             'error': None,
#             'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
#             'steps': [],
#             'test_logs': []
#         }
#
#         try:
#             formatted_url = URLValidator.format_url(url)
#
#             # Check for PaloAlto blocked URL first
#             if "ceas.sase.responses.es.oneadp.com" in url:
#                 log_step('SKIP', f"Launching URL Is ==> {formatted_url}")
#                 try:
#                     driver = URLValidator.create_web_driver()
#                     driver.get(formatted_url)
#                     time.sleep(2)
#                     screenshot_base64 = URLValidator.capture_screenshot_base64(driver)
#                 except Exception as e:
#                     logging.warning(f"Error capturing screenshot for blocked URL: {str(e)}")
#
#                 result.update({
#                     'status': 'Skip',
#                     'error': "Access blocked by security policy",
#                     'category': 'PALOALTO_BLOCK',
#                     'screenshot_base64': screenshot_base64,
#                     'test_logs': test_logs
#                 })
#                 return result
#
#             if not URLValidator.is_valid_url(url):
#                 log_step('Failed', f"Invalid URL format: {url}")
#                 result.update({
#                     'status': 'Failed',
#                     'error': "Invalid URL format",
#                     'category': 'ERROR',
#                     'test_logs': test_logs
#                 })
#                 return result
#
#             driver = URLValidator.create_web_driver()
#
#             try:
#                 # First log entry for launching URL
#                 log_step('INFO', f"Launching URL Is ==> {formatted_url}")
#
#                 # Attempt to load the URL
#                 driver.get(formatted_url)
#
#                 # Always capture these logs regardless of status
#                 page_title = driver.title
#                 redirected_url = driver.current_url
#                 load_time = (time.time() - start_time) * 1000
#
#                 # Log all mandatory fields
#                 log_step('INFO', f"Title Of The Page Is ==> {page_title}")
#                 log_step('INFO', f"Redirected URL Is ==> {redirected_url}")
#                 log_step('INFO', f"Time Taken to Launch Application - {formatted_url} ==> {load_time:.2f} Milliseconds")
#
#                 # Always attempt to capture screenshot
#                 try:
#                     time.sleep(2)
#                     screenshot_base64 = URLValidator.capture_screenshot_base64(driver)
#                     if screenshot_base64:
#                         result['screenshot_base64'] = screenshot_base64
#                 except Exception as screenshot_error:
#                     logging.warning(f"Screenshot capture failed: {str(screenshot_error)}")
#
#                 # Check page status - This determines the final status
#                 status_result = URLValidator.check_url_status(driver)
#                 result.update(status_result)
#                 result['load_time'] = round(load_time, 2)
#
#                 # Try to load page completely - but don't override the status from check_url_status
#                 if result['status'] != 'Skip':
#                     page_loaded = URLValidator.check_page_loaded(
#                         driver,
#                         timeout=config.get("page_load_timeout", 30),
#                         retries=config.get("max_retries", 2),
#                         retry_delay=config.get("retry_delay", 2)
#                     )
#
#                     # Only log the page load status, don't change the result status
#                     if not page_loaded:
#                         log_step('Failed', "Page failed to load completely")
#                     else:
#                         log_step('Success', "Page loaded successfully")
#
#             except Exception as e:
#                 result.update({
#                     'status': 'Failed',
#                     'error': str(e),
#                     'category': 'ERROR',
#                 })
#
#             # Update test logs based on final status
#             result['test_logs'] = test_logs
#
#         except Exception as e:
#             result.update({
#                 'status': 'Failed',
#                 'error': str(e),
#                 'category': 'ERROR',
#             })
#
#         finally:
#             # Ensure all required logs are present
#             result['test_logs'] = test_logs
#             if driver:
#                 if not result.get('screenshot_base64'):
#                     try:
#                         screenshot_base64 = URLValidator.capture_screenshot_base64(driver)
#                         if screenshot_base64:
#                             result['screenshot_base64'] = screenshot_base64
#                     except Exception:
#                         pass
#                     try:
#                         driver.quit()
#                     except Exception:
#                         pass
#
#         return result
#
#     @staticmethod
#     def is_valid_url(url):
#         """Validate URL format"""
#         try:
#             result = urlparse(URLValidator.format_url(url))
#             return all([result.scheme, result.netloc])
#         except Exception:
#             return False
#
#     @staticmethod
#     def format_url(url):
#         """Format URL with proper protocol"""
#         if not url:
#             return ""
#         url = url.strip()
#         if not url.startswith(("https://", "http://")):
#             return f"https://{url}"
#         return url
#
#     # @staticmethod
#     # def check_url_status(driver):
#     #     """Check URL status with enhanced categorization"""
#     #     try:
#     #         current_url = driver.current_url.lower()
#     #         page_source = driver.page_source.lower() if driver.page_source else ""
#     #         page_title = driver.title if driver.title else ""
#     #
#     #         navigation_start = driver.execute_script("return window.performance.timing.navigationStart")
#     #         response_end = driver.execute_script("return window.performance.timing.responseEnd")
#     #         duration = response_end - navigation_start
#     #
#     #         # Check for PaloAlto block first
#     #         if "ceas.sase.responses.es.oneadp.com" in current_url or \
#     #                 "ceas.sase.responses.es.oneadp.com" in page_source:
#     #             return {
#     #                 'status': 'Skip',
#     #                 'error': 'Access blocked by security policy',
#     #                 'category': 'PALOALTO_BLOCK',
#     #                 'load_time': duration,
#     #                 'steps': [{
#     #                     'status': 'SKIP',
#     #                     'message': f'URL is blocked by security policy\nTitle: {page_title}\nURL: {current_url}',
#     #                     'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #                 }]
#     #             }
#     #
#     #         # Define error categories
#     #         error_categories = {
#     #             'CERTIFICATE_ERROR': {
#     #                 'status': 'Failed',
#     #                 'error_codes': ['ERR_CERT_DATE_INVALID', 'ERR_CERT_AUTHORITY_INVALID',
#     #                                 'SEC_ERROR_EXPIRED_CERTIFICATE'],
#     #                 'messages': ['certificate has expired', 'ssl certificate error', 'certificate is not trusted']
#     #             },
#     #             'CONNECTION_ERROR': {
#     #                 'status': 'Failed',
#     #                 'error_codes': ['ERR_CONNECTION_TIMED_OUT', 'ERR_CONNECTION_REFUSED', 'ERR_NAME_NOT_RESOLVED'],
#     #                 'messages': ['connection timed out', 'connection was refused', 'name not resolved']
#     #             },
#     #             'SERVER_ERROR': {
#     #                 'status': 'Failed',
#     #                 'error_codes': ['500', '502', '503', '504'],
#     #                 'messages': ['500 internal server error', '502 bad gateway', '503 service unavailable',
#     #                              '504 gateway timeout']
#     #             },
#     #             'DNS_ERROR': {
#     #                 'status': 'Failed',
#     #                 'error_codes': ['ERR_NAME_NOT_RESOLVED', 'DNS_PROBE_FINISHED_NXDOMAIN'],
#     #                 'messages': ['dns lookup failed', 'dns error', 'domain not found']
#     #             },
#     #             'ACCESS_DENIED': {
#     #                 'status': 'Failed',
#     #                 'error_codes': ['ERR_ACCESS_DENIED', '403'],
#     #                 'messages': ['access denied', '403 forbidden', '401 unauthorized']
#     #             },
#     #             'NOT_FOUND': {
#     #                 'status': 'Failed',
#     #                 'error_codes': ['404'],
#     #                 'messages': ['404 not found', 'page not found', 'content not found']
#     #             },
#     #             'SLOW_RESPONSE': {
#     #                 'status': 'Warning',
#     #                 'error_codes': ['SLOW_RESPONSE'],
#     #                 'messages': ['slow server response', 'taking longer than expected']
#     #             },
#     #             'SECURITY_WARNING': {
#     #                 'status': 'Warning',
#     #                 'error_codes': ['SEC_WARNING'],
#     #                 'messages': ['security warning', 'potential security risk']
#     #             },
#     #             'ACCESS_BLOCKED': {
#     #                 'status': 'Skip',
#     #                 'error_codes': ['ACCESS_BLOCKED', 'SECURITY_BLOCK'],
#     #                 'messages': [
#     #                     'web - access blocked',
#     #                     'access blocked by security policy',
#     #                     'blocked by administrator',
#     #                     'this site has been blocked',
#     #                     'content filtered'
#     #                 ]
#     #             },
#     #             'PROXY_BLOCK': {
#     #                 'status': 'Skip',
#     #                 'error_codes': ['PROXY_BLOCK'],
#     #                 'messages': ['proxy block', 'blocked by network policy']
#     #             }
#     #         }
#     #
#     #         try:
#     #             error_elements = driver.find_elements(By.CLASS_NAME, "error-code")
#     #             if error_elements:
#     #                 error_text = " ".join([elem.text.lower() for elem in error_elements])
#     #
#     #                 for category, details in error_categories.items():
#     #                     if any(code.lower() in error_text for code in details['error_codes']):
#     #                         return {
#     #                             'status': details['status'],
#     #                             'error': f'{category}: {error_text}',
#     #                             'category': category,
#     #                             'load_time': duration,
#     #                             'steps': [{
#     #                                 'status': details['status'].upper(),
#     #                                 'message': f'{category} detected\nError Code: {error_text}\nTitle: {page_title}\nURL: {current_url}',
#     #                                 'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #                             }]
#     #                         }
#     #
#     #             page_text = page_source + " " + page_title.lower()
#     #             for category, details in error_categories.items():
#     #                 if any(msg in page_text for msg in details['messages']):
#     #                     return {
#     #                         'status': details['status'],
#     #                         'error': f'{category} - {next((msg for msg in details["messages"] if msg in page_text), "")}',
#     #                         'category': category,
#     #                         'load_time': duration,
#     #                         'steps': [{
#     #                             'status': details['status'].upper(),
#     #                             'message': f'{category} detected\nTitle: {page_title}\nURL: {current_url}',
#     #                             'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #                         }]
#     #                     }
#     #
#     #             return {
#     #                 'status': 'Success',
#     #                 'error': None,
#     #                 'category': 'SUCCESS',
#     #                 'load_time': duration,
#     #                 'steps': [{
#     #                     'status': 'PASS',
#     #                     'message': f'Page loaded successfully\nTitle: {page_title}\nURL: {current_url}',
#     #                     'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #                 }]
#     #             }
#     #
#     #         except Exception as e:
#     #             logging.error(f"Error checking page content: {str(e)}")
#     #             return {
#     #                 'status': 'Failed',
#     #                 'error': str(e),
#     #                 'category': 'CONTENT_CHECK_ERROR',
#     #                 'load_time': duration,
#     #                 'steps': [{
#     #                     'status': 'FAIL',
#     #                     'message': f'Error checking page content: {str(e)}\nTitle: {page_title}\nURL: {current_url}',
#     #                     'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #                 }]
#     #             }
#     #
#     #     except Exception as e:
#     #         logging.error(f"Error in check_url_status: {str(e)}")
#     #         return {
#     #             'status': 'Failed',
#     #             'error': str(e),
#     #             'category': 'SYSTEM_ERROR',
#     #             'load_time': 0,
#     #             'steps': [{
#     #                 'status': 'FAIL',
#     #                 'message': f'System error during status check: {str(e)}',
#     #                 'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #             }]
#     #         }
#
#     @staticmethod
#     def format_validation_result(url, status, title, redirected_url, duration, error=None, category=None):
#         """Format validation result in a consistent way for HTML reporting"""
#         timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#
#         # Base result structure
#         result = {
#             'url': url,
#             'status': status,
#             'category': category if category else ('ERROR' if status == 'Failed' else 'SUCCESS'),
#             'load_time': round(duration, 2),
#             'test_logs': []
#         }
#
#         # For Skip status, only show URL and blocked message
#         if status == 'Skip':
#             log_entries = [
#                 {
#                     'timestamp': timestamp,
#                     'level': 'SKIP',
#                     'message': f"Launching URL Is ==> {url}"
#                 },
#                 {
#                     'timestamp': timestamp,
#                     'level': 'SKIP',
#                     'message': f"Page Is Blocked By PaloAlto"
#                 }
#             ]
#         else:
#             # For all other statuses (Success, Failed, Warning), show all fields
#             log_entries = [
#                 {
#                     'timestamp': timestamp,
#                     'level': status.upper(),
#                     'message': f"Launching URL Is ==> {url}"
#                 },
#                 {
#                     'timestamp': timestamp,
#                     'level': 'INFO',
#                     'message': f"Title Of The Page Is ==> {title}"
#                 },
#                 {
#                     'timestamp': timestamp,
#                     'level': 'INFO',
#                     'message': f"Redirected URL Is ==> {redirected_url}"
#                 },
#                 {
#                     'timestamp': timestamp,
#                     'level': 'INFO',
#                     'message': f"Time Taken to Launch Application - {url} is ==> {duration:.2f} Milliseconds"
#                 }
#             ]
#
#             # Add error message if present for non-Skip statuses
#             if error:
#                 log_entries.append({
#                     'timestamp': timestamp,
#                     'level': 'ERROR',
#                     'message': error
#                 })
#
#         result['test_logs'] = log_entries
#         return result
#
#     # @staticmethod
#     # def check_url_status(driver):
#     #     """Check URL status with grouped error conditions"""
#     #     try:
#     #         current_url = driver.current_url.lower()
#     #         page_source = driver.page_source.lower() if driver.page_source else ""
#     #         page_title = driver.title if driver.title else ""
#     #
#     #         navigation_start = driver.execute_script("return window.performance.timing.navigationStart")
#     #         response_end = driver.execute_script("return window.performance.timing.responseEnd")
#     #         duration = response_end - navigation_start
#     #
#     #         # SKIP CONDITIONS
#     #         if any(pattern in page_source or pattern in current_url for pattern in [
#     #             "ceas.sase.responses.es.oneadp.com",
#     #             "web - access blocked",
#     #             "access blocked by security policy",
#     #             "blocked by administrator",
#     #             "this site has been blocked",
#     #             "content filtered"
#     #         ]):
#     #             logging.info("Page Is Blocked By PaloAlto")
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Skip',
#     #                 title="",
#     #                 redirected_url=current_url,
#     #                 duration=duration,
#     #                 error='Page Is Blocked By PaloAlto',
#     #                 category='PALOALTO_BLOCK'
#     #             )
#     #
#     #         # FAILED CONDITIONS
#     #         if any(pattern in page_source for pattern in [
#     #             "web look up failed",
#     #             "this site can't be reached",
#     #             "dns lookup failed",
#     #             "dns_probe_finished_nxdomain",
#     #             "access denied",
#     #             "403",
#     #             "404",
#     #             "page not found",
#     #             "500",
#     #             "internal server error",
#     #             "502",
#     #             "bad gateway",
#     #             "503",
#     #             "service unavailable",
#     #             "504",
#     #             "gateway timeout",
#     #             "certificate error",
#     #             "ssl error",
#     #             "connection refused",
#     #             "connection timed out",
#     #             "err_connection_refused",
#     #             "err_connection_timed_out"
#     #         ]):
#     #             # Determine specific error category and message
#     #             if "dns" in page_source or "web look up failed" in page_source:
#     #                 error_msg = "DNS lookup failed"
#     #                 category = "DNS_ERROR"
#     #             elif "403" in page_source or "access denied" in page_source:
#     #                 error_msg = "403 Forbidden"
#     #                 category = "ACCESS_DENIED"
#     #             elif "404" in page_source or "page not found" in page_source:
#     #                 error_msg = "404 Page Not Found"
#     #                 category = "NOT_FOUND"
#     #             elif any(code in page_source for code in ["500", "502", "503", "504"]):
#     #                 error_msg = f"{next(code for code in ['500', '502', '503', '504'] if code in page_source)} Server Error"
#     #                 category = "SERVER_ERROR"
#     #             elif "certificate" in page_source or "ssl error" in page_source:
#     #                 error_msg = "SSL/Certificate Error"
#     #                 category = "CERTIFICATE_ERROR"
#     #             else:
#     #                 error_msg = "Connection Error"
#     #                 category = "CONNECTION_ERROR"
#     #
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Failed',
#     #                 title=page_title,
#     #                 redirected_url=driver.current_url,
#     #                 duration=duration,
#     #                 error=error_msg,
#     #                 category=category
#     #             )
#     #
#     #         # WARNING CONDITIONS
#     #         if any(pattern in page_source for pattern in [
#     #             "security warning",
#     #             "potential security risk",
#     #             "this site isn't secure",
#     #             "mixed content warning",
#     #             "connection is not fully secure"
#     #         ]) or duration > 10000:  # Slow response (>10s)
#     #             error_msg = "Security warning detected" if "security" in page_source else "Slow server response"
#     #             category = "SECURITY_WARNING" if "security" in page_source else "SLOW_RESPONSE"
#     #
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Warning',
#     #                 title=page_title,
#     #                 redirected_url=driver.current_url,
#     #                 duration=duration,
#     #                 error=error_msg,
#     #                 category=category
#     #             )
#     #
#     #         # SUCCESS CONDITION
#     #         if driver.execute_script("return document.readyState") == "complete":
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Success',
#     #                 title=page_title,
#     #                 redirected_url=driver.current_url,
#     #                 duration=duration
#     #             )
#     #
#     #         # Default failure case
#     #         return URLValidator.format_validation_result(
#     #             url=current_url,
#     #             status='Failed',
#     #             title=page_title,
#     #             redirected_url=driver.current_url,
#     #             duration=duration,
#     #             error='Page failed to load properly',
#     #             category='ERROR'
#     #         )
#     #
#     #     except Exception as e:
#     #         logging.error(f"Error in check_url_status: {str(e)}")
#     #         return URLValidator.format_validation_result(
#     #             url=current_url if 'current_url' in locals() else "Unknown URL",
#     #             status='Failed',
#     #             title="Error",
#     #             redirected_url="",
#     #             duration=0,
#     #             error=str(e),
#     #             category='SYSTEM_ERROR'
#     #         )
#
#     # @staticmethod
#     # def check_url_status(driver):
#     #     """Check URL status with grouped error conditions"""
#     #     try:
#     #         current_url = driver.current_url.lower()
#     #         page_source = driver.page_source.lower() if driver.page_source else ""
#     #         page_title = driver.title if driver.title else ""
#     #
#     #         navigation_start = driver.execute_script("return window.performance.timing.navigationStart")
#     #         response_end = driver.execute_script("return window.performance.timing.responseEnd")
#     #         duration = response_end - navigation_start
#     #
#     #         # First check - PaloAlto block
#     #         if any(pattern in page_source or pattern in current_url for pattern in [
#     #             "ceas.sase.responses.es.oneadp.com",
#     #             "web - access blocked",
#     #             "access blocked by security policy",
#     #             "blocked by administrator",
#     #             "this site has been blocked",
#     #             "content filtered"
#     #         ]):
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Skip',
#     #                 title=page_title,
#     #                 redirected_url=current_url,
#     #                 duration=duration,
#     #                 error='Page Is Blocked By PaloAlto',
#     #                 category='PALOALTO_BLOCK'
#     #             )
#     #
#     #         # Second check - Failed conditions with expanded patterns
#     #         failed_patterns = [
#     #             # Domain specific failures
#     #             "httpstat.us/403",
#     #             "msess.gov.pt",
#     #             "expired.badssl.com",
#     #             # HTTP error codes
#     #             "403", "forbidden", "access denied",
#     #             "404", "not found",
#     #             "400", "bad request",
#     #             "401", "unauthorized",
#     #             "500", "internal server error",
#     #             "502", "bad gateway",
#     #             "503", "service unavailable",
#     #             "504", "gateway timeout",
#     #             # Common error messages
#     #             "web look up failed",
#     #             "this site can't be reached",
#     #             "dns lookup failed",
#     #             "dns_probe_finished_nxdomain",
#     #             "err_connection_refused",
#     #             "connection refused",
#     #             "connection timed out",
#     #             "name not resolved",
#     #             # SSL/Security errors
#     #             "certificate error",
#     #             "certificate has expired",
#     #             "ssl certificate error",
#     #             "cert_date_invalid",
#     #             "certificate is not trusted",
#     #             "expired certificate"
#     #         ]
#     #
#     #         if any(pattern in page_source or pattern in current_url for pattern in failed_patterns):
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Failed',
#     #                 title=page_title,
#     #                 redirected_url=driver.current_url,
#     #                 duration=duration,
#     #                 error='Connection/Server/Certificate Error',
#     #                 category='ERROR'
#     #             )
#     #
#     #         # Third check - Warning conditions for slow response and security warnings
#     #         warning_domains = ['slow.badssl.com']
#     #         warning_patterns = [
#     #             "security warning",
#     #             "potential security risk",
#     #             "this site isn't secure",
#     #             "mixed content warning",
#     #             "connection is not fully secure"
#     #         ]
#     #
#     #         if (any(domain in current_url for domain in warning_domains) or
#     #                 any(pattern in page_source for pattern in warning_patterns) or
#     #                 duration > 3000):  # 3 seconds threshold for slow responses
#     #
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Warning',
#     #                 title=page_title,
#     #                 redirected_url=driver.current_url,
#     #                 duration=duration,
#     #                 error='Warning: Performance/Security concerns detected',
#     #                 category='WARNING'
#     #             )
#     #
#     #         # Success condition - only if all above checks pass
#     #         return URLValidator.format_validation_result(
#     #             url=current_url,
#     #             status='Success',
#     #             title=page_title,
#     #             redirected_url=driver.current_url,
#     #             duration=duration
#     #         )
#     #
#     #     except Exception as e:
#     #         logging.error(f"Error in check_url_status: {str(e)}")
#     #         return URLValidator.format_validation_result(
#     #             url=current_url if 'current_url' in locals() else "Unknown URL",
#     #             status='Failed',
#     #             title="Error",
#     #             redirected_url="",
#     #             duration=0,
#     #             error=str(e),
#     #             category='SYSTEM_ERROR'
#     #         )
#
#     # @staticmethod
#     # def check_url_status(driver):
#     #     """Check URL status with clearly separated status checks for Skip, Failed, Warning, and Success"""
#     #     try:
#     #         current_url = driver.current_url.lower()
#     #         page_source = driver.page_source.lower() if driver.page_source else ""
#     #         page_title = driver.title if driver.title else ""
#     #
#     #         navigation_start = driver.execute_script("return window.performance.timing.navigationStart")
#     #         response_end = driver.execute_script("return window.performance.timing.responseEnd")
#     #         duration = response_end - navigation_start
#     #
#     #         # 1. First Check: SKIP status
#     #         if (
#     #                 "ceas.sase.responses.es.oneadp.com" in current_url
#     #                 # or "web - access blocked" in page_source or
#     #                 # "access blocked by security policy" in page_source or
#     #                 # "blocked by administrator" in page_source or
#     #                 # "this site has been blocked" in page_source or
#     #                 # "content filtered" in page_source
#     #         ):
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Warning',
#     #                 title=page_title,
#     #                 redirected_url=current_url,
#     #                 duration=duration,
#     #                 error='Page Is Blocked By PaloAlto',
#     #                 category='BLOCKED-PAGE'
#     #             )
#     #         elif(
#     #             "This page isn't working" in page_source
#     #         ):
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Fail',
#     #                 title=page_title,
#     #                 redirected_url=current_url,
#     #                 duration=duration,
#     #                 error='Page is failing',
#     #                 category='ERROR'
#     #             )
#     #         elif(
#     #             "web - access blocked" in page_source
#     #         ):
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Skip',
#     #                 title=page_title,
#     #                 redirected_url=current_url,
#     #                 duration=duration,
#     #                 error='Web Access Blocked',
#     #                 category='SKIP'
#     #             )
#     #         else:
#     #             return URLValidator.format_validation_result(
#     #                 url=current_url,
#     #                 status='Pass',
#     #                 title=page_title,
#     #                 redirected_url=current_url,
#     #                 duration=duration,
#     #                 error='Page is displaying correctly',
#     #                 category='PASSED'
#     #             )
#     #     finally:
#     #         print("Page is loading correctly")
#     #
#     #     #     # 2. Second Check: FAILED status
#     #     #     if (
#     #     #             # HTTP error codes
#     #     #             "/403" in current_url or
#     #     #             "/404" in current_url or
#     #     #             "/500" in current_url or
#     #     #             "/502" in current_url or
#     #     #             "/503" in current_url or
#     #     #             "/504" in current_url or
#     #     #             # Specific domains known to fail
#     #     #             "msess.gov.pt" in current_url or
#     #     #             "expired.badssl" in current_url or
#     #     #             # Error messages in page source
#     #     #             any(error in page_source for error in [
#     #     #                 "403 forbidden",
#     #     #                 "404 not found",
#     #     #                 "500 internal server error",
#     #     #                 "502 bad gateway",
#     #     #                 "503 service unavailable",
#     #     #                 "504 gateway timeout",
#     #     #                 "certificate has expired",
#     #     #                 "ssl certificate error",
#     #     #                 "cert_date_invalid",
#     #     #                 "certificate is not trusted",
#     #     #                 "unable to connect",
#     #     #                 "connection refused",
#     #     #                 "dns_probe_finished",
#     #     #                 "err_cert_authority_invalid",
#     #     #                 "err_cert_date_invalid",
#     #     #                 "this site can't be reached",
#     #     #                 "net::err_connection_refused",
#     #     #                 "net::err_cert_authority_invalid",
#     #     #                 "net::err_cert_date_invalid"
#     #     #             ])
#     #     #     ):
#     #     #         error_message = (
#     #     #             "HTTP Error" if any(
#     #     #                 code in current_url for code in ["/403", "/404", "/500", "/502", "/503", "/504"])
#     #     #             else "SSL Certificate Error" if "expired.badssl" in current_url or any(
#     #     #                 cert in page_source for cert in ["certificate", "ssl", "cert"])
#     #     #             else "Connection Error"
#     #     #         )
#     #     #         return URLValidator.format_validation_result(
#     #     #             url=current_url,
#     #     #             status='Failed',
#     #     #             title=page_title,
#     #     #             redirected_url=driver.current_url,
#     #     #             duration=duration,
#     #     #             error=error_message,
#     #     #             category='ERROR'
#     #     #         )
#     #     #
#     #     #     # 3. Third Check: WARNING status
#     #     #     if (
#     #     #             # Slow loading domains
#     #     #             "slow.badssl.com" in current_url or
#     #     #             # Performance threshold
#     #     #             duration > 3000 or
#     #     #             # Security warnings
#     #     #             any(warning in page_source for warning in [
#     #     #                 "security warning",
#     #     #                 "potential security risk",
#     #     #                 "this site isn't secure",
#     #     #                 "mixed content warning",
#     #     #                 "connection is not fully secure"
#     #     #             ])
#     #     #     ):
#     #     #         warning_message = "Performance Warning: Slow Response" if duration > 3000 else "Security Warning: Potential Risks Detected"
#     #     #         return URLValidator.format_validation_result(
#     #     #             url=current_url,
#     #     #             status='Warning',  # This will now show as WARNING instead of PASS
#     #     #             title=page_title,
#     #     #             redirected_url=driver.current_url,
#     #     #             duration=duration,
#     #     #             error=warning_message,
#     #     #             category='WARNING'
#     #     #         )
#     #     #
#     #     #     # 4. Fourth Check: SUCCESS status
#     #     #     if driver.execute_script("return document.readyState") == "complete":
#     #     #         return URLValidator.format_validation_result(
#     #     #             url=current_url,
#     #     #             status='Success',
#     #     #             title=page_title,
#     #     #             redirected_url=driver.current_url,
#     #     #             duration=duration
#     #     #         )
#     #     #
#     #     #     # Default case if none of the above conditions match
#     #     #     return URLValidator.format_validation_result(
#     #     #         url=current_url,
#     #     #         status='Failed',
#     #     #         title=page_title,
#     #     #         redirected_url=driver.current_url,
#     #     #         duration=duration,
#     #     #         error='Page failed to load completely',
#     #     #         category='ERROR'
#     #     #     )
#     #     #
#     #     # except Exception as e:
#     #     #     logging.error(f"Error in check_url_status: {str(e)}")
#     #     #     return URLValidator.format_validation_result(
#     #     #         url=current_url if 'current_url' in locals() else "Unknown URL",
#     #     #         status='Failed',
#     #     #         title="Error",
#     #     #         redirected_url="",
#     #     #         duration=0,
#     #     #         error=str(e),
#     #     #         category='SYSTEM_ERROR'
#     #     #     )
#
#     @staticmethod
#     def check_url_status(driver):
#         """Check URL status with clearly separated status checks"""
#         try:
#             current_url = driver.current_url.lower()
#             page_source = driver.page_source.lower() if driver.page_source else ""
#             page_title = driver.title if driver.title else ""
#
#             navigation_start = driver.execute_script("return window.performance.timing.navigationStart")
#             response_end = driver.execute_script("return window.performance.timing.responseEnd")
#             duration = response_end - navigation_start
#
#             # Check conditions in order of priority
#             if "ceas.sase.responses.es.oneadp.com" in current_url:
#                 return ReportHandler.format_validation_result(
#                     url=current_url,
#                     status='Warning',
#                     title=page_title,
#                     redirected_url=current_url,
#                     duration=duration,
#                     error='Page Is Blocked By PaloAlto',
#                     category='BLOCKED-PAGE'
#                 )
#             elif "This page isn't working" in page_source:
#                 return ReportHandler.format_validation_result(
#                     url=current_url,
#                     status='Failed',
#                     title=page_title,
#                     redirected_url=current_url,
#                     duration=duration,
#                     error='Page is failing',
#                     category='ERROR'
#                 )
#             elif "web - access blocked" in page_source:
#                 return ReportHandler.format_validation_result(
#                     url=current_url,
#                     status='Skip',
#                     title=page_title,
#                     redirected_url=current_url,
#                     duration=duration,
#                     error='Web Access Blocked',
#                     category='SKIP'
#                 )
#             else:
#                 return ReportHandler.format_validation_result(
#                     url=current_url,
#                     status='Success',
#                     title=page_title,
#                     redirected_url=current_url,
#                     duration=duration,
#                     error='Page is displaying correctly',
#                     category='PASSED'
#                 )
#
#         except Exception as e:
#             logging.error(f"Error in check_url_status: {str(e)}")
#             return ReportHandler.format_validation_result(
#                 url=current_url if 'current_url' in locals() else "Unknown URL",
#                 status='Failed',
#                 title="Error",
#                 redirected_url="",
#                 duration=0,
#                 error=str(e),
#                 category='SYSTEM_ERROR'
#             )
#
#     @staticmethod
#     def check_page_loaded(driver, timeout=30, retries=2, retry_delay=2):
#         """Check if page is loaded with retries"""
#         current_url = driver.current_url
#         print(f"\n{Fore.CYAN}► Attempting to load page: {current_url}{Style.RESET_ALL}")
#         print(f"{Fore.YELLOW}⚙ Settings: Timeout={timeout}s, Retries={retries}, Delay={retry_delay}s{Style.RESET_ALL}")
#
#         for attempt in range(retries):
#             try:
#                 print(f"\n{Fore.MAGENTA}● Attempt {attempt + 1}/{retries}:{Style.RESET_ALL}")
#                 print(f"{Fore.CYAN}⌛ Checking document ready state...{Style.RESET_ALL}")
#
#                 wait = WebDriverWait(driver, timeout)
#                 wait.until(lambda driver: driver.execute_script("return document.readyState") == "complete")
#                 print(f"{Fore.GREEN}✓ Document ready state is complete{Style.RESET_ALL}")
#
#                 print(f"{Fore.CYAN}⌛ Checking jQuery status...{Style.RESET_ALL}")
#                 jquery_check = """
#                         return (typeof jQuery !== 'undefined') ?
#                             jQuery.active == 0 : true
#                     """
#                 wait.until(lambda driver: driver.execute_script(jquery_check))
#                 print(f"{Fore.GREEN}✓ jQuery ajax calls completed{Style.RESET_ALL}")
#
#                 print(f"\n{Fore.GREEN}✓ Success: Page loaded successfully!{Style.RESET_ALL}")
#                 print(f"{Fore.GREEN}✓ Total attempts needed: {attempt + 1}/{retries}{Style.RESET_ALL}")
#                 return True
#
#             except Exception:
#                 if attempt < retries - 1:
#                     print(
#                         f"{Fore.YELLOW}⚠ Timeout occurred. Waiting {retry_delay} seconds before retry...{Style.RESET_ALL}")
#                     time.sleep(retry_delay)
#                 else:
#                     print(f"{Fore.RED}✗ Final attempt failed{Style.RESET_ALL}")
#                 continue
#
#         print(f"\n{Fore.RED}✗ Failed: Page did not load after all retry attempts{Style.RESET_ALL}")
#         print(f"{Fore.RED}✗ URL that failed to load: {current_url}{Style.RESET_ALL}")
#         return False
#
#     @staticmethod
#     def capture_screenshot_base64(driver):
#         """Capture screenshot and convert to base64"""
#         try:
#             driver.execute_script("document.body.style.overflow = 'hidden';")
#             screenshot = driver.get_screenshot_as_png()
#             driver.execute_script("document.body.style.overflow = '';")
#             return base64.b64encode(screenshot).decode('utf-8')
#         except Exception as e:
#             logging.error(f"Error capturing screenshot: {str(e)}")
#             return None
#
#     @staticmethod
#     def read_excel_urls(excel_path, sheet_name):
#         """Read URLs from Excel file"""
#         try:
#             if not os.path.exists(excel_path):
#                 raise FileNotFoundError(f"Excel file not found: {excel_path}")
#
#             df = pd.read_excel(excel_path, sheet_name=sheet_name)
#             if 'URL' not in df.columns:
#                 raise ValueError("Excel file must contain a 'URL' column")
#
#             urls = df['URL'].dropna().str.strip().tolist()
#             logging.info(f"Read {len(urls)} URLs from Excel")
#             return [(url, idx + 2) for idx, url in enumerate(urls)]
#
#         except Exception as e:
#             logging.error(f"Error reading Excel file: {str(e)}")
#             raise
#
#     @staticmethod
#     def backup_previous_excel():
#         """Backup previous Excel report"""
#         try:
#             paths = URLValidator.get_project_paths()
#             excel_dir = paths["output_excel"]
#             backup_dir = paths["backup_excel"]
#
#             os.makedirs(backup_dir, exist_ok=True)
#
#             if not os.path.exists(excel_dir) or not os.listdir(excel_dir):
#                 return
#
#             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#
#             for filename in os.listdir(excel_dir):
#                 if filename.endswith('.xlsx'):
#                     src_path = os.path.join(excel_dir, filename)
#                     backup_filename = f"Previous-{filename.replace('.xlsx', '')}_{timestamp}.xlsx"
#                     dst_path = os.path.join(backup_dir, backup_filename)
#
#                     try:
#                         if os.path.exists(src_path) and os.access(src_path, os.R_OK):
#                             if os.path.exists(dst_path):
#                                 os.remove(dst_path)
#                             shutil.copy2(src_path, dst_path)
#                             os.remove(src_path)
#                             logging.info(f"Backed up Excel report: {dst_path}")
#                         else:
#                             logging.warning(f"Source file not accessible or missing: {src_path}")
#
#                     except Exception as e:
#                         logging.error(f"Error backing up {src_path}: {str(e)}")
#                         continue
#
#         except Exception as e:
#             logging.error(f"Error during Excel backup: {str(e)}")
#
#     @staticmethod
#     def generate_excel_report(results):
#         """Generate Excel report from results"""
#         try:
#             paths = URLValidator.get_project_paths()
#             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#
#             URLValidator.backup_previous_excel()
#
#             # Updated data structure to include error messages
#             data = [{
#                 'URL': result['url'],
#                 'Status': result['status'],
#                 'Load Time (ms)': round(result.get('load_time', 0), 2),
#                 'Error/Warning': result.get('error', '')
#             } for result in results]
#
#             df = pd.DataFrame(data)
#             os.makedirs(paths["output_excel"], exist_ok=True)
#             output_path = os.path.join(paths["output_excel"], "URL-Validation-Report.xlsx")
#
#             try:
#                 # Apply conditional formatting for different statuses
#                 with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
#                     df.to_excel(writer, index=False)
#                     workbook = writer.book
#                     worksheet = writer.sheets['Sheet1']
#
#                     # Define styles for different statuses
#                     success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
#                     warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
#                     fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
#                     skip_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
#
#                     # Apply conditional formatting
#                     for row in range(2, len(df) + 2):  # Skip header row
#                         status_cell = worksheet.cell(row=row, column=2)  # Status column
#                         if status_cell.value == 'Success':
#                             status_cell.fill = success_fill
#                         elif status_cell.value == 'Warning':
#                             status_cell.fill = warning_fill
#                         elif status_cell.value == 'Failed':
#                             status_cell.fill = fail_fill
#                         elif status_cell.value == 'Skip':
#                             status_cell.fill = skip_fill
#
#             except PermissionError:
#                 alt_output_path = os.path.join(paths["output_excel"], f"URL-Validation-Report_{timestamp}.xlsx")
#                 df.to_excel(alt_output_path, index=False, engine='openpyxl')
#                 output_path = alt_output_path
#                 logging.warning(f"Could not save to default filename, saved as: {alt_output_path}")
#
#             logging.info(f"Excel report generated: {output_path}")
#             return output_path
#
#         except Exception as e:
#             logging.error(f"Error generating Excel report: {str(e)}")
#             raise
#
#     #@staticmethod
#     # def format_validation_result(url, status, title, redirected_url, duration, error=None, category=None):
#     #     """Format validation result in a consistent way"""
#     #     result = {
#     #         'url': url,
#     #         'status': status,
#     #         'category': category if category else ('ERROR' if status == 'Failed' else 'SUCCESS'),
#     #         'load_time': round(duration, 2),
#     #         'test_logs': []
#     #     }
#     #
#     #     # Add standard log entries
#     #     timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#     #     log_entries = [
#     #         {
#     #             'timestamp': timestamp,
#     #             'level': status.upper(),
#     #             'message': f"Launching URL Is ==> {url}"
#     #         },
#     #         {
#     #             'timestamp': timestamp,
#     #             'level': 'INFO',
#     #             'message': f"Title Of The Page Is ==> {title}"
#     #         },
#     #         {
#     #             'timestamp': timestamp,
#     #             'level': 'INFO',
#     #             'message': f"Redirected URL Is ==> {redirected_url}"
#     #         },
#     #         {
#     #             'timestamp': timestamp,
#     #             'level': 'INFO',
#     #             'message': f"Time Taken to Launch Application - {url} is ==> {duration:.2f} Milliseconds"
#     #         }
#     #     ]
#         # # Add error message if present
#         # if error:
#         #     log_entries.append({
#         #         'timestamp': timestamp,
#         #         'level': 'ERROR',
#         #         'message': error
#         #     })
#         #
#         # result['test_logs'] = log_entries
#         # return result
#
# class ValidationRunner:
#     def __init__(self):
#         """Initialize the validation runner"""
#         colorama.init(autoreset=True)
#         self.start_time = datetime.now()
#         self.log_path = URLValidator.setup_logging()
#         self.setup_console_output()
#         logging.info("ValidationRunner initialized")
#         logging.info(f"Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}")
#         logging.info(f"Log File: {self.log_path}")
#
#     def setup_console_output(self):
#         """Setup console output"""
#         os.system('cls' if os.name == 'nt' else 'clear')
#         console_handler = logging.StreamHandler(sys.stdout)
#         console_handler.setLevel(logging.INFO)
#         formatter = logging.Formatter(
#             f'{Fore.YELLOW}%(asctime)s{Style.RESET_ALL} - '
#             f'{Fore.CYAN}%(levelname)s{Style.RESET_ALL} - '
#             f'%(message)s',
#             datefmt='%H:%M:%S'
#         )
#         console_handler.setFormatter(formatter)
#
#         root_logger = logging.getLogger()
#         for handler in root_logger.handlers[:]:
#             if isinstance(handler, logging.StreamHandler) and handler.stream == sys.stdout:
#                 root_logger.removeHandler(handler)
#
#         root_logger.addHandler(console_handler)
#
#     def print_header(self):
#         """Print header information"""
#         print(f"\n{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}")
#         print(f"{Fore.GREEN}URL Validation Process{Style.RESET_ALL}".center(80))
#         print(f"{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}\n")
#         print(f"{Fore.YELLOW}Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}{Style.RESET_ALL}")
#         print(f"{Fore.YELLOW}Log File: {self.log_path}{Style.RESET_ALL}\n")
#
#     def print_summary(self, results):
#         """Print validation summary including warnings"""
#         end_time = datetime.now()
#         duration = end_time - self.start_time
#
#         # Count different statuses
#         success_count = sum(1 for r in results if r['status'] == 'Success')
#         warning_count = sum(1 for r in results if r['status'] == 'Warning')
#         skip_count = sum(1 for r in results if r['status'] == 'Skip')
#         fail_count = sum(1 for r in results if r['status'] == 'Failed')
#
#         # Calculate success rate excluding skipped URLs but counting warnings
#         total_excluding_skipped = len(results) - skip_count
#         success_rate = (success_count / total_excluding_skipped * 100) if total_excluding_skipped > 0 else 0
#
#         print(f"\n{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}")
#         print(f"{Fore.GREEN}Validation Summary{Style.RESET_ALL}".center(80))
#         print(f"{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}\n")
#
#         print(f"{Fore.YELLOW}Duration: {duration}{Style.RESET_ALL}")
#         print(f"{Fore.YELLOW}Total URLs Processed: {len(results)}{Style.RESET_ALL}")
#         print(f"{Fore.GREEN}✓ Successful: {success_count}{Style.RESET_ALL}")
#         print(f"{Fore.RED}✗ Failed: {fail_count}{Style.RESET_ALL}")
#         print(f"{Fore.YELLOW}⚠ Warnings: {warning_count}{Style.RESET_ALL}")
#         print(f"{Fore.BLUE}○ Skipped: {skip_count}{Style.RESET_ALL}")
#         print(f"{Fore.YELLOW}Success Rate: {success_rate:.2f}%{Style.RESET_ALL}\n")
#
#         # Show URLs by status
#         if fail_count > 0:
#             print(f"{Fore.RED}Failed URLs:{Style.RESET_ALL}")
#             for result in results:
#                 if result['status'] == 'Failed':
#                     print(f"{Fore.RED}✗ {result['url']}: {result.get('error', 'Unknown error')}{Style.RESET_ALL}")
#
#         if warning_count > 0:
#             print(f"\n{Fore.YELLOW}Warning URLs:{Style.RESET_ALL}")
#             for result in results:
#                 if result['status'] == 'Warning':
#                     print(
#                         f"{Fore.YELLOW}⚠ {result['url']}: {result.get('error', 'Performance/Security Warning')}{Style.RESET_ALL}")
#
#         if skip_count > 0:
#             print(f"\n{Fore.BLUE}Skipped URLs:{Style.RESET_ALL}")
#             for result in results:
#                 if result['status'] == 'Skip':
#                     print(f"{Fore.BLUE}○ {result['url']}: {result.get('error', 'Unknown error')}{Style.RESET_ALL}")
#
#     def validate_urls(self):
#         """Run the URL validation process"""
#         try:
#             self.print_header()
#             config = URLValidator.load_config()
#             paths = URLValidator.get_project_paths()
#
#             print(f"\n{Fore.CYAN}Reading URLs from Excel...{Style.RESET_ALL}")
#             urls = URLValidator.read_excel_urls(paths["excel_path"], config["sheet_name"])
#
#             if not urls:
#                 print(f"\n{Fore.RED}✗ Error: No URLs found in Excel file{Style.RESET_ALL}")
#                 return
#
#             total_urls = len(urls)
#             print(f"\n{Fore.GREEN}✓ Found {total_urls} URLs to process{Style.RESET_ALL}")
#
#             results = []
#             with tqdm(total=total_urls, desc="Processing URLs",
#                       bar_format="{l_bar}%s{bar}%s{r_bar}" % (Fore.GREEN, Style.RESET_ALL)) as pbar:
#
#                 for i, (url, row_number) in enumerate(urls, 1):
#                     try:
#                         result = URLValidator.process_url(url, row_number, config)
#                         results.append(result)
#
#                         # Updated status color logic to include warnings
#                         if result['status'] == 'Success':
#                             status_color = Fore.GREEN
#                             status_symbol = "✓"
#                         elif result['status'] == 'Warning':
#                             status_color = Fore.YELLOW
#                             status_symbol = "⚠"
#                         elif result['status'] == 'Skip':
#                             status_color = Fore.BLUE
#                             status_symbol = "○"
#                         else:
#                             status_color = Fore.RED
#                             status_symbol = "✗"
#
#                         pbar.set_postfix_str(
#                             f"{status_color}{status_symbol} {result['status']}{Style.RESET_ALL}",
#                             refresh=True
#                         )
#                         pbar.update(1)
#
#                         if i < total_urls:
#                             time.sleep(config.get("wait_between_urls", 2))
#
#                     except Exception as e:
#                         print(f"\n{Fore.RED}✗ Error processing URL {url}: {str(e)}{Style.RESET_ALL}")
#                         results.append({
#                             'url': url,
#                             'status': 'Failed',
#                             'error': str(e),
#                             'load_time': 0,
#                             'steps': []
#                         })
#                         pbar.update(1)
#
#             print(f"\n{Fore.CYAN}Generating reports...{Style.RESET_ALL}")
#
#             try:
#                 excel_report = URLValidator.generate_excel_report(results)
#                 print(f"{Fore.GREEN}✓ Excel report generated: {excel_report}{Style.RESET_ALL}")
#
#                 from utils.report_handler import ReportHandler
#                 html_report = ReportHandler.generate_detailed_html_report(results, paths["reports"])
#                 print(f"{Fore.GREEN}✓ HTML report generated: {html_report}{Style.RESET_ALL}")
#
#             except Exception as e:
#                 print(f"{Fore.RED}✗ Error generating reports: {str(e)}{Style.RESET_ALL}")
#
#             self.print_summary(results)
#
#         except Exception as e:
#             print(f"\n{Fore.RED}✗ Validation process failed: {str(e)}{Style.RESET_ALL}")
#             print(f"\n{Fore.RED}Stack trace:{Style.RESET_ALL}")
#             print(traceback.format_exc())
#
#         finally:
#             colorama.deinit()
#
# def main():
#     """Main entry point"""
#     try:
#         print(f"\n{Fore.CYAN}Starting URL validation process...{Style.RESET_ALL}")
#         runner = ValidationRunner()
#         runner.validate_urls()
#         print(f"\n{Fore.GREEN}✓ URL validation process completed successfully{Style.RESET_ALL}")
#
#     except Exception as e:
#         print(f"\n{Fore.RED}✗ Fatal error in main process: {str(e)}{Style.RESET_ALL}")
#         print(f"\n{Fore.RED}Stack trace:{Style.RESET_ALL}")
#         print(traceback.format_exc())
#         sys.exit(1)
#
#     finally:
#         print(f"\n{Fore.YELLOW}Process finished. Check reports for detailed results.{Style.RESET_ALL}")
#
# if __name__ == "__main__":
#     main()


import os
import shutil
import sys
import json
import time
import logging
from logging.handlers import RotatingFileHandler
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse
import traceback
from tqdm import tqdm
import colorama
from colorama import Fore, Style
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, WebDriverException
from openpyxl.styles import PatternFill
import base64
from report_handler import ReportHandler

class URLValidator:
    @staticmethod
    def load_config():
        """Load configuration from config.json"""
        try:
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config_path = os.path.join(project_root, 'config.json')

            if not os.path.exists(config_path):
                raise FileNotFoundError(f"Configuration file not found: {config_path}")

            with open(config_path, 'r') as f:
                config = json.load(f)

            logging.info("Configuration loaded successfully")
            return config
        except Exception as e:
            logging.error(f"Error loading configuration: {str(e)}")
            raise

    @staticmethod
    def get_project_paths():
        """Get project paths"""
        try:
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            config = URLValidator.load_config()

            paths = {
                "project_root": project_root,
                "excel_path": os.path.join(project_root, config.get("excel_path", "resources/links.xlsx")),
                "chrome_driver": os.path.join(project_root,
                                            config.get("chrome_driver_path", "resources/drivers/chromedriver.exe")),
                "reports": os.path.join(project_root, "reports", "validation-reports"),
                "logs": os.path.join(project_root, "logs"),
                "logs_current": os.path.join(project_root, "logs", "Current Logs"),
                "logs_backup": os.path.join(project_root, "logs", "Backup Logs"),
                "output_excel": os.path.join(project_root, "reports", "excel-reports"),
                "backup_excel": os.path.join(project_root, "reports", "backup-excel")
            }

            logging.info("Project paths initialized")
            return paths
        except Exception as e:
            logging.error(f"Error getting project paths: {str(e)}")
            raise

    @staticmethod
    def setup_logging():
        """Initialize logging configuration"""
        try:
            paths = URLValidator.get_project_paths()
            os.makedirs(paths["logs_current"], exist_ok=True)
            os.makedirs(paths["logs_backup"], exist_ok=True)
            current_log_file = os.path.join(paths["logs_current"], "Current-Validation-Logs.log")

            if os.path.exists(current_log_file):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup_filename = f"Previous-Validation-Logs-{timestamp}.log"
                backup_path = os.path.join(paths["logs_backup"], backup_filename)

                try:
                    shutil.copy2(current_log_file, backup_path)
                    open(current_log_file, 'w').close()
                except Exception as e:
                    print(f"Error backing up log file: {str(e)}")

            root_logger = logging.getLogger()
            root_logger.setLevel(logging.DEBUG)
            root_logger.handlers.clear()

            file_formatter = logging.Formatter(
                '%(asctime)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            console_formatter = logging.Formatter(
                '%(asctime)s - %(levelname)s - %(message)s',
                datefmt='%H:%M:%S'
            )

            file_handler = RotatingFileHandler(
                current_log_file,
                maxBytes=10*1024*1024,
                backupCount=5,
                encoding='utf-8'
            )
            file_handler.setLevel(logging.DEBUG)
            file_handler.setFormatter(file_formatter)
            root_logger.addHandler(file_handler)

            console_handler = logging.StreamHandler(sys.stdout)
            console_handler.setLevel(logging.INFO)
            console_handler.setFormatter(console_formatter)
            root_logger.addHandler(console_handler)

            logging.info("Logging setup completed")
            return current_log_file

        except Exception as e:
            print(f"Error setting up logging: {str(e)}")
            raise

    @staticmethod
    def ensure_directories():
        """Create necessary directories"""
        try:
            paths = URLValidator.get_project_paths()
            for path in paths.values():
                if path and not os.path.exists(path):
                    os.makedirs(path)
                    logging.info(f"Created directory: {path}")

            os.makedirs(paths["logs_current"], exist_ok=True)
            os.makedirs(paths["logs_backup"], exist_ok=True)
            logging.info("Directory structure verified")
        except Exception as e:
            logging.error(f"Error creating directories: {str(e)}")
            raise

    @staticmethod
    def create_web_driver():
        """Create and configure Chrome WebDriver"""
        try:
            chrome_options = Options()
            chrome_options.add_argument("--headless")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--ignore-certificate-errors")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--remote-allow-origins=*")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-extensions")
            chrome_options.add_argument("--disable-popup-blocking")
            chrome_options.set_capability("acceptInsecureCerts", True)

            os.environ["no_proxy"] = "*"
            os.environ["HTTP_PROXY"] = ""
            os.environ["HTTPS_PROXY"] = ""

            paths = URLValidator.get_project_paths()
            config = URLValidator.load_config()

            if not os.path.exists(paths["chrome_driver"]):
                raise FileNotFoundError(f"ChromeDriver not found at: {paths['chrome_driver']}")

            service = Service(executable_path=paths["chrome_driver"])
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(config.get("page_load_timeout", 60))
            driver.set_window_size(1920, 1080)

            logging.info("WebDriver created successfully")
            return driver

        except Exception as e:
            logging.error(f"Error creating WebDriver: {str(e)}")
            raise

    @staticmethod
    def process_url(url, row_number, config):
        """Process a single URL and return results with improved screenshot capture for all URLs"""
        start_time = time.time()
        driver = None
        test_logs = []

        def log_step(level, message, symbol="●"):
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            if "Launching URL Is ==>" in message:
                if result.get('status') == 'Success':
                    level = 'PASS'
                elif result.get('status') == 'Failed':
                    level = 'FAIL'
                elif result.get('status') == 'Warning':
                    level = 'WARNING'
                elif result.get('status') == 'Skip':
                    level = 'SKIP'

            log_entry = {
                'timestamp': timestamp,
                'level': level,
                'message': message
            }
            test_logs.append(log_entry)

            if level in ['ERROR', 'FAIL', 'Failed']:
                print(f"{Fore.RED}✗ [{timestamp}] {message}{Style.RESET_ALL}")
            elif level in ['WARNING', 'Warning', 'warning']:
                print(f"{Fore.YELLOW}⚠ [{timestamp}] {message}{Style.RESET_ALL}")
            elif level in ['SUCCESS', 'PASS']:
                print(f"{Fore.GREEN}✓ [{timestamp}] {message}{Style.RESET_ALL}")
            elif level in ['SKIP', 'Skip']:
                print(f"{Fore.BLUE}○ [{timestamp}] {message}{Style.RESET_ALL}")
            else:
                print(f"{Fore.CYAN}{symbol} [{timestamp}] {message}{Style.RESET_ALL}")

        def capture_screenshot(driver, attempt_number=""):
            """Helper function to capture screenshot with logging"""
            try:
                time.sleep(2)  # Wait for page stability
                driver.execute_script("document.body.style.overflow = 'hidden';")
                screenshot = driver.get_screenshot_as_png()
                driver.execute_script("document.body.style.overflow = '';")
                if screenshot:
                    log_step('SUCCESS', f"Screenshot captured successfully{attempt_number}")
                    return base64.b64encode(screenshot).decode('utf-8')
                log_step('WARNING', f"Screenshot capture returned no data{attempt_number}")
                return None
            except Exception as e:
                log_step('WARNING', f"Screenshot capture failed{attempt_number}: {str(e)}")
                return None

        result = {
            'url': url,
            'status': None,
            'category': 'ERROR',
            'load_time': 0,
            'screenshot_base64': None,
            'error': None,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'steps': [],
            'test_logs': []
        }

        try:
            formatted_url = URLValidator.format_url(url)
            driver = URLValidator.create_web_driver()
            driver.set_window_size(1920, 1080)  # Ensure consistent screenshot size

            try:
                # Initial page load for all URLs
                log_step('INFO', f"Launching URL Is ==> {formatted_url}")
                driver.get(formatted_url)
                time.sleep(2)  # Initial wait for page load

                # First screenshot attempt
                screenshot_base64 = capture_screenshot(driver, " (first attempt)")
                if screenshot_base64:
                    result['screenshot_base64'] = screenshot_base64

                # Special handling for PaloAlto blocked URLs
                if "ceas.sase.responses.es.oneadp.com" in url:
                    if not result.get('screenshot_base64'):
                        # Second attempt for blocked URLs
                        screenshot_base64 = capture_screenshot(driver, " (retry for blocked URL)")
                        if screenshot_base64:
                            result['screenshot_base64'] = screenshot_base64

                    result.update({
                        'status': 'Warning',
                        'error': "Page Is Blocked By PaloAlto",
                        'category': 'WARNING',
                        'test_logs': test_logs
                    })
                    return result

                if not URLValidator.is_valid_url(url):
                    result.update({
                        'status': 'Failed',
                        'error': "Invalid URL format",
                        'category': 'ERROR',
                        'test_logs': test_logs
                    })
                    return result

                # Normal URL processing
                page_title = driver.title
                redirected_url = driver.current_url
                load_time = (time.time() - start_time) * 1000

                # Take screenshot after successful page load
                if not result.get('screenshot_base64'):
                    screenshot_base64 = capture_screenshot(driver, " (after successful load)")
                    if screenshot_base64:
                        result['screenshot_base64'] = screenshot_base64

                log_step('INFO', f"Title Of The Page Is ==> {page_title}")
                log_step('INFO', f"Redirected URL Is ==> {redirected_url}")
                log_step('INFO', f"Time Taken to Launch Application - {formatted_url} ==> {load_time:.2f} Milliseconds")

                # Check page status
                status_result = URLValidator.check_url_status(driver)
                result.update(status_result)
                result['load_time'] = round(load_time, 2)

                # Final screenshot attempt if still missing
                if not result.get('screenshot_base64'):
                    screenshot_base64 = capture_screenshot(driver, " (final attempt)")
                    if screenshot_base64:
                        result['screenshot_base64'] = screenshot_base64

            except Exception as e:
                if "ceas.sase.responses.es.oneadp.com" in url:
                    result.update({
                        'status': 'Warning',
                        'error': "Page Is Blocked By PaloAlto",
                        'category': 'WARNING',
                    })
                else:
                    result.update({
                        'status': 'Failed',
                        'error': str(e),
                        'category': 'ERROR',
                    })

                # Try to capture error state screenshot
                if not result.get('screenshot_base64'):
                    screenshot_base64 = capture_screenshot(driver, " (error state)")
                    if screenshot_base64:
                        result['screenshot_base64'] = screenshot_base64

        except Exception as e:
            result.update({
                'status': 'Failed' if "ceas.sase.responses.es.oneadp.com" not in url else 'Warning',
                'error': str(e),
                'category': 'ERROR' if "ceas.sase.responses.es.oneadp.com" not in url else 'WARNING',
            })

        finally:
            result['test_logs'] = test_logs
            if driver:
                # Last chance screenshot capture
                if not result.get('screenshot_base64'):
                    screenshot_base64 = capture_screenshot(driver, " (final cleanup)")
                    if screenshot_base64:
                        result['screenshot_base64'] = screenshot_base64
                try:
                    driver.quit()
                except Exception:
                    pass

            return result

    @staticmethod
    def is_valid_url(url):
        """Validate URL format"""
        try:
            result = urlparse(URLValidator.format_url(url))
            return all([result.scheme, result.netloc])
        except:
            return False

    @staticmethod
    def format_url(url):
        """Format URL with proper protocol"""
        if not url:
            return ""
        url = url.strip()
        if not url.startswith(("https://", "http://")):
            return f"https://{url}"
        return url

    @staticmethod
    def check_url_status(driver):
        """Check URL status with updated warning logic"""
        try:
            current_url = driver.current_url.lower()
            page_source = driver.page_source.lower() if driver.page_source else ""
            page_title = driver.title if driver.title else ""

            navigation_start = driver.execute_script("return window.performance.timing.navigationStart")
            response_end = driver.execute_script("return window.performance.timing.responseEnd")
            duration = response_end - navigation_start

            # Check for PaloAlto URLs and related errors first
            if ("ceas.sase.responses.es.oneadp.com" in current_url or
                    "ceas.sase.responses.es.oneadp.com" in page_source or
                    ("net::err_name_not_resolved" in page_source and "ceas.sase.responses.es.oneadp.com" in str(
                        current_url))):
                return ReportHandler.format_validation_result(
                    url=current_url,
                    status='Warning',
                    title=page_title,
                    redirected_url=current_url,
                    duration=duration,
                    error='Page Is Blocked By PaloAlto',
                    category='WARNING'
                )
            elif "This page isn't working" in page_source:
                return ReportHandler.format_validation_result(
                    url=current_url,
                    status='Failed',
                    title=page_title,
                    redirected_url=current_url,
                    duration=duration,
                    error='Page is failing',
                    category='ERROR'
                )
            elif "web - access blocked" in page_source:
                return ReportHandler.format_validation_result(
                    url=current_url,
                    status='Skip',
                    title=page_title,
                    redirected_url=current_url,
                    duration=duration,
                    error='Web Access Blocked',
                    category='SKIP'
                )
            else:
                return ReportHandler.format_validation_result(
                    url=current_url,
                    status='Success',
                    title=page_title,
                    redirected_url=current_url,
                    duration=duration,
                    error=None,
                    category='SUCCESS'
                )

        except Exception as e:
            logging.error(f"Error in check_url_status: {str(e)}")
            return ReportHandler.format_validation_result(
                url=current_url if 'current_url' in locals() else "Unknown URL",
                status='Failed',
                title="Error",
                redirected_url="",
                duration=0,
                error=str(e),
                category='SYSTEM_ERROR'
            )

    @staticmethod
    def capture_screenshot_base64(driver):
        """Capture screenshot and convert to base64"""
        try:
            driver.execute_script("document.body.style.overflow = 'hidden';")
            screenshot = driver.get_screenshot_as_png()
            driver.execute_script("document.body.style.overflow = '';")
            return base64.b64encode(screenshot).decode('utf-8')
        except Exception as e:
            logging.error(f"Error capturing screenshot: {str(e)}")
            return None

    @staticmethod
    def read_excel_urls(excel_path, sheet_name):
        """Read URLs from Excel file"""
        try:
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"Excel file not found: {excel_path}")

            df = pd.read_excel(excel_path, sheet_name=sheet_name)
            if 'URL' not in df.columns:
                raise ValueError("Excel file must contain a 'URL' column")

            urls = df['URL'].dropna().str.strip().tolist()
            logging.info(f"Read {len(urls)} URLs from Excel")
            return [(url, idx + 2) for idx, url in enumerate(urls)]

        except Exception as e:
            logging.error(f"Error reading Excel file: {str(e)}")
            raise

    @staticmethod
    def backup_previous_excel():
        """Backup previous Excel report"""
        try:
            paths = URLValidator.get_project_paths()
            excel_dir = paths["output_excel"]
            backup_dir = paths["backup_excel"]

            os.makedirs(backup_dir, exist_ok=True)

            if not os.path.exists(excel_dir) or not os.listdir(excel_dir):
                return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            for filename in os.listdir(excel_dir):
                if filename.endswith('.xlsx'):
                    src_path = os.path.join(excel_dir, filename)
                    backup_filename = f"Previous-{filename.replace('.xlsx', '')}_{timestamp}.xlsx"
                    dst_path = os.path.join(backup_dir, backup_filename)

                    try:
                        if os.path.exists(src_path) and os.access(src_path, os.R_OK):
                            if os.path.exists(dst_path):
                                os.remove(dst_path)
                            shutil.copy2(src_path, dst_path)
                            os.remove(src_path)
                            logging.info(f"Backed up Excel report: {dst_path}")
                        else:
                            logging.warning(f"Source file not accessible or missing: {src_path}")

                    except Exception as e:
                        logging.error(f"Error backing up {src_path}: {str(e)}")
                        continue

        except Exception as e:
            logging.error(f"Error during Excel backup: {str(e)}")

    @staticmethod
    def generate_excel_report(results):
        """Generate Excel report from results"""
        try:
            paths = URLValidator.get_project_paths()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            URLValidator.backup_previous_excel()

            data = [{
                'URL': result['url'],
                'Status': result['status'],
                'Load Time (ms)': round(result.get('load_time', 0), 2),
                'Error/Warning': result.get('error', '')
            } for result in results]

            df = pd.DataFrame(data)
            os.makedirs(paths["output_excel"], exist_ok=True)
            output_path = os.path.join(paths["output_excel"], "URL-Validation-Report.xlsx")

            try:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']

                    # Define styles for different statuses
                    success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    skip_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

                    # Apply conditional formatting
                    for row in range(2, len(df) + 2):  # Skip header row
                        status_cell = worksheet.cell(row=row, column=2)  # Status column
                        if status_cell.value == 'Success':
                            status_cell.fill = success_fill
                        elif status_cell.value == 'Warning':
                            status_cell.fill = warning_fill
                        elif status_cell.value == 'Failed':
                            status_cell.fill = fail_fill
                        elif status_cell.value == 'Skip':
                            status_cell.fill = skip_fill

            except PermissionError:
                alt_output_path = os.path.join(paths["output_excel"], f"URL-Validation-Report_{timestamp}.xlsx")
                df.to_excel(alt_output_path, index=False, engine='openpyxl')
                output_path = alt_output_path
                logging.warning(f"Could not save to default filename, saved as: {alt_output_path}")

            logging.info(f"Excel report generated: {output_path}")
            return output_path

        except Exception as e:
            logging.error(f"Error generating Excel report: {str(e)}")
            raise


class ValidationRunner:
    def __init__(self):
        """Initialize the validation runner"""
        colorama.init(autoreset=True)
        self.start_time = datetime.now()
        self.log_path = URLValidator.setup_logging()
        self.setup_console_output()
        #print(f"Add a blank line before starting")
        print(f"{Fore.YELLOW}--------------------------")
        logging.info("ValidationRunner initialized")
        logging.info(f"Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        logging.info(f"Log File: {self.log_path}")

    def setup_console_output(self):
        """Setup console output"""
        os.system('cls' if os.name == 'nt' else 'clear')
        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        formatter = logging.Formatter(
            f'{Fore.YELLOW}%(asctime)s{Style.RESET_ALL} - '
            f'{Fore.CYAN}%(levelname)s{Style.RESET_ALL} - '
            f'%(message)s',
            datefmt='%H:%M:%S'
        )
        console_handler.setFormatter(formatter)

        root_logger = logging.getLogger()
        for handler in root_logger.handlers[:]:
            if isinstance(handler, logging.StreamHandler) and handler.stream == sys.stdout:
                root_logger.removeHandler(handler)
        root_logger.addHandler(console_handler)

    def print_header(self):
        """Print header information"""
        print(f"\n{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}URL Validation Process{Style.RESET_ALL}".center(80))
        print(f"{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}\n")
        print(f"{Fore.YELLOW}Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Log File: {self.log_path}{Style.RESET_ALL}\n")

    def print_summary(self, results):
        """Print validation summary"""
        end_time = datetime.now()
        duration = end_time - self.start_time

        success_count = sum(1 for r in results if r['status'] == 'Success')
        warning_count = sum(1 for r in results if r['status'] == 'Warning')
        skip_count = sum(1 for r in results if r['status'] == 'Skip')
        fail_count = sum(1 for r in results if r['status'] == 'Failed')

        total_excluding_skipped = len(results) - skip_count
        success_rate = (success_count / total_excluding_skipped * 100) if total_excluding_skipped > 0 else 0

        print(f"\n{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}Validation Summary{Style.RESET_ALL}".center(80))
        print(f"{Fore.CYAN}{'=' * 80}{Style.RESET_ALL}\n")

        print(f"{Fore.YELLOW}Duration: {duration}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Total URLs Processed: {len(results)}{Style.RESET_ALL}")
        print(f"{Fore.GREEN}✓ Successful: {success_count}{Style.RESET_ALL}")
        print(f"{Fore.RED}✗ Failed: {fail_count}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}⚠ Warnings: {warning_count}{Style.RESET_ALL}")
        print(f"{Fore.BLUE}○ Skipped: {skip_count}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}Success Rate: {success_rate:.2f}%{Style.RESET_ALL}\n")

        if fail_count > 0:
            print(f"{Fore.RED}Failed URLs:{Style.RESET_ALL}")
            for result in results:
                if result['status'] == 'Failed':
                    print(f"{Fore.RED}✗ {result['url']}: {result.get('error', 'Unknown error')}{Style.RESET_ALL}")

        if warning_count > 0:
            print(f"\n{Fore.YELLOW}Warning URLs:{Style.RESET_ALL}")
            for result in results:
                if result['status'] == 'Warning':
                    print(
                        f"{Fore.YELLOW}⚠ {result['url']}: {result.get('error', 'Performance/Security Warning')}{Style.RESET_ALL}")

        if skip_count > 0:
            print(f"\n{Fore.BLUE}Skipped URLs:{Style.RESET_ALL}")
            for result in results:
                if result['status'] == 'Skip':
                    print(f"{Fore.BLUE}○ {result['url']}: {result.get('error', 'Unknown error')}{Style.RESET_ALL}")

    def validate_urls(self):
        """Run the URL validation process"""
        try:
            self.print_header()
            config = URLValidator.load_config()
            paths = URLValidator.get_project_paths()

            print(f"\n{Fore.CYAN}Reading URLs from Excel...{Style.RESET_ALL}")
            urls = URLValidator.read_excel_urls(paths["excel_path"], config["sheet_name"])

            if not urls:
                print(f"\n{Fore.RED}✗ Error: No URLs found in Excel file{Style.RESET_ALL}")
                return

            total_urls = len(urls)
            print(f"\n{Fore.GREEN}✓ Found {total_urls} URLs to process{Style.RESET_ALL}")
            # print("Add blank line before progress bar")
            print(f"\n{Fore.GREEN}-----------------------------")

            results = []
            with tqdm(total=total_urls, desc="Processing URLs",
                      bar_format="{l_bar}%s{bar}%s{r_bar}" % (Fore.GREEN, Style.RESET_ALL)) as pbar:

                print("newline after the progress bar is initialized")

                for i, (url, row_number) in enumerate(urls, 1):
                    try:
                        result = URLValidator.process_url(url, row_number, config)
                        results.append(result)

                        status_symbol = {
                            'Success': ('✓', Fore.GREEN),
                            'Warning': ('⚠', Fore.YELLOW),
                            'Skip': ('○', Fore.BLUE),
                            'Failed': ('✗', Fore.RED)
                        }.get(result['status'], ('✗', Fore.RED))

                        pbar.set_postfix_str(
                            f"{status_symbol[1]}{status_symbol[0]} {result['status']}{Style.RESET_ALL}",
                            refresh=True
                        )
                        pbar.update(1)

                        if i < total_urls:
                            time.sleep(config.get("wait_between_urls", 2))

                    except Exception as e:
                        print(f"\n{Fore.RED}✗ Error processing URL {url}: {str(e)}{Style.RESET_ALL}")
                        results.append({
                            'url': url,
                            'status': 'Failed',
                            'error': str(e),
                            'load_time': 0,
                            'steps': []
                        })
                        pbar.update(1)

                    #print(f"\n newline after each URL is processed")
                    print(f"\n{Fore.GREEN}=========================newline after each URL is processed")

            print(f"\n{Fore.CYAN}Generating reports...{Style.RESET_ALL}")

            try:
                excel_report = URLValidator.generate_excel_report(results)
                print(f"{Fore.GREEN}✓ Excel report generated: {excel_report}{Style.RESET_ALL}")

                html_report = ReportHandler.generate_detailed_html_report(results, paths["reports"])
                print(f"{Fore.GREEN}✓ HTML report generated: {html_report}{Style.RESET_ALL}")

            except Exception as e:
                print(f"{Fore.RED}✗ Error generating reports: {str(e)}{Style.RESET_ALL}")

            self.print_summary(results)

        except Exception as e:
            print(f"\n{Fore.RED}✗ Validation process failed: {str(e)}{Style.RESET_ALL}")
            print(f"\n{Fore.RED}Stack trace:{Style.RESET_ALL}")
            print(traceback.format_exc())

        finally:
            colorama.deinit()


def main():
    """Main entry point"""
    try:
        print(f"\n{Fore.CYAN}Starting URL validation process...{Style.RESET_ALL}")
        runner = ValidationRunner()
        runner.validate_urls()
        print(f"\n{Fore.GREEN}✓ URL validation process completed successfully{Style.RESET_ALL}")

    except Exception as e:
        print(f"\n{Fore.RED}✗ Fatal error in main process: {str(e)}{Style.RESET_ALL}")
        print(f"\n{Fore.RED}Stack trace:{Style.RESET_ALL}")
        print(traceback.format_exc())
        sys.exit(1)

    finally:
        print(f"\n{Fore.YELLOW}Process finished. Check reports for detailed results.{Style.RESET_ALL}")


if __name__ == "__main__":
    main()